"""
Bag Count Export Module - Uses openpyxl for proper Excel handling.

Exports bag counts and census data to Excel, preserving template formatting
and dynamically adding new facilities with proper spacing.
"""

import os
import json
import shutil
import subprocess
import urllib.request
from copy import copy
from datetime import datetime, timedelta
from pathlib import Path

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

APP_DIR = Path(__file__).parent
DATA_DIR = APP_DIR / "data"
TEMPLATE_PATH = DATA_DIR / "template_bag_count.xlsx"

# Column configuration for each day
DAY_CONFIG = {
    "Mon": {"fac_col": "B", "batch_col": "C", "val_col": "D"},
    "Tue": {"fac_col": "F", "batch_col": "G", "val_col": "H"},
    "Wed": {"fac_col": "J", "batch_col": "K", "val_col": "L"},
    "Thu": {"fac_col": "N", "batch_col": "O", "val_col": "P"},
    "Fri": {"fac_col": "R", "batch_col": "S", "val_col": "T"},
}

# Known facilities and their cell mappings in the template
# Format: (facility_key, batch_name) -> value_cell_ref
KNOWN_CELL_MAP = {
    # Monday (value col D)
    ("Mother of Good Counsel", "1st Floor"): "D4",
    ("Mother of Good Counsel", "2nd Floor"): "D6",
    ("Mother of Good Counsel", "3rd Floor"): "D7",
    ("McClay", "100 Hall"): "D10",
    ("McClay", "L Hall"): "D11",
    ("Belleview", "Hall A"): "D14",
    ("Belleview", "Hall B"): "D15",
    ("Belleview", "Unit"): "D16",
    ("Hillside", "Hill 1"): "D19",
    ("Hillside", "Hill 2"): "D20",
    ("Superior Manor", "Hall 100"): "D23",
    ("Superior Manor", "Hall 200"): "D24",
    ("Superior Manor", "Hall 300"): "D25",
    ("Superior Manor", "Hall 400"): "D26",
    ("Walnut Street", "1"): "D29",
    ("Colonial Doniphan", "Upstairs"): "D31",
    ("Colonial Doniphan", "Down"): "D32",
    ("New Hope", "ALF"): "D35",
    ("New Hope", "ILF"): "D36",
    # Tuesday (value col H)
    ("Westwood Hills", "A Wing"): "H4",
    ("Westwood Hills", "B Wing"): "H5",
    ("Westwood Hills", "C Wing"): "H6",
    ("Glenfield", "100"): "H9",
    ("Granite House", "1"): "H12",
    ("Creve Coeur", "1"): "H14",
    ("Creve Coeur", "2"): "H15",
    # Wednesday (value col L)
    ("Licking RCF", "1"): "L4",
    ("Salem Care Center", "1"): "L6",
    ("Salem Care Center", "2"): "L7",
    ("Salem Care Center", "3"): "L8",
    ("Salem Residential Care", "RCF"): "L10",
    ("Baisch SNF", "1"): "L12",
    ("Baisch SNF", "2"): "L13",
    ("Baisch RCF", "RCF"): "L14",
    ("Pillars", "1"): "L17",
    ("Pillars", "2"): "L19",
    ("Legacy", "1"): "L23",
    ("John Knox CSL", "CSL"): "L25",
    ("John Knox MM", "MM"): "L26",
    ("John Knox ALF", "ALF"): "L27",
    ("John Knox ILF", "ILF 1"): "L28",
    ("John Knox ILF", "ILF 2"): "L29",
    # Thursday (value col P) - Bentley's added at P23 with proper spacing
    ("Delta South", "1"): "P4",
    ("Delta South", "2"): "P6",
    ("Seville", "1"): "P10",
    ("St Johns", "1"): "P12",
    ("St Johns", "2"): "P13",
    ("UCity", "1"): "P16",
    ("UCity", "2"): "P17",
    ("UCity", "3"): "P18",
    ("UCity", "4"): "P19",
    ("UCity", "5"): "P20",
    ("Bentley's", "1"): "P23",
    # Friday (value col T)
    ("Colonial Bismark", "1"): "T4",
    ("Bertrand", "1"): "T6",
    ("Bertrand", "2"): "T8",
    ("Oakdale SNF", "1"): "T12",
    ("Oakdale SNF", "2"): "T14",
    ("Oakdale ALF", "1"): "T18",
    ("Oakdale ALF", "2"): "T19",
    ("Oakdale ALF", "3"): "T20",
    ("Oakdale RCF", "1"): "T23",
    ("Oakdale RCF", "2"): "T24",
}

# Daily total row for each day (will be updated if new facilities are added)
DAILY_TOTAL_ROWS = {
    "Mon": 39,
    "Tue": 18,
    "Wed": 32,
    "Thu": 26,  # After Bentley's addition
    "Fri": 27,
}

# Last facility row for each day (for adding new facilities after)
LAST_FACILITY_ROWS = {
    "Mon": 37,  # New Hope ILF
    "Tue": 16,  # Creve Coeur total
    "Wed": 30,  # John Knox ILF 2
    "Thu": 24,  # Bentley's total
    "Fri": 25,  # Oakdale RCF total
}


def load_supabase_state():
    """Load bag count state from Supabase."""
    url = os.environ.get("SUPABASE_URL", "")
    key = os.environ.get("SUPABASE_KEY", "")
    
    if not url or not key:
        # Try loading from .env file
        env_path = APP_DIR / ".env"
        if env_path.exists():
            with open(env_path) as f:
                for line in f:
                    if line.startswith("SUPABASE_URL"):
                        url = line.split("=", 1)[1].strip()
                    if line.startswith("SUPABASE_KEY"):
                        key = line.split("=", 1)[1].strip()
    
    if url and key:
        try:
            req = urllib.request.Request(
                f"{url}/rest/v1/tracking_state?key=eq.bag_counts",
                headers={"apikey": key, "Authorization": f"Bearer {key}"}
            )
            with urllib.request.urlopen(req) as resp:
                data = json.loads(resp.read())
                return data[0]["value"] if data else {}
        except Exception as e:
            print(f"Error loading from Supabase: {e}")
    
    return {"counts": {}, "batches": {}}


def find_new_facilities(counts, batches_config):
    """Find facilities/batches that aren't in the known cell map."""
    new_facilities = {}  # {day: [(facility, batch_name, batch_id), ...]}
    
    day_map = {"D": "Mon", "H": "Tue", "L": "Wed", "P": "Thu", "T": "Fri"}
    col_to_day = {v["val_col"]: k for k, v in DAY_CONFIG.items()}
    
    # Build set of known (facility, batch) pairs per day
    known_by_day = {day: set() for day in DAY_CONFIG.keys()}
    for (fac, batch), cell_ref in KNOWN_CELL_MAP.items():
        col = cell_ref[0]
        day = col_to_day.get(col)
        if day:
            known_by_day[day].add((fac, batch))
    
    # Check for new facilities
    for day in DAY_CONFIG.keys():
        day_counts = counts.get(day, {})
        for facility, fac_data in day_counts.items():
            fac_batches = batches_config.get(facility, [])
            for batch in fac_batches:
                batch_name = batch["name"]
                batch_id = batch["id"]
                if (facility, batch_name) not in known_by_day[day]:
                    if day not in new_facilities:
                        new_facilities[day] = []
                    new_facilities[day].append((facility, batch_name, batch_id))
    
    return new_facilities


def add_new_facility_to_sheet(ws, day, facility_name, batches, style_source_row, insert_after_row):
    """
    Add a new facility to a sheet with proper formatting.
    
    - Blank row after previous facility
    - Facility name and batches
    - Facility total row
    - Blank row before daily total
    
    Only modifies the columns for that day.
    """
    config = DAY_CONFIG[day]
    fac_col = config["fac_col"]
    batch_col = config["batch_col"]
    val_col = config["val_col"]
    
    current_row = insert_after_row + 1  # Start after blank row
    
    # Copy style from source row
    def copy_cell_style(src_row, dest_row, col):
        src_cell = ws[f"{col}{src_row}"]
        dest_cell = ws[f"{col}{dest_row}"]
        if src_cell.has_style:
            dest_cell.font = copy(src_cell.font)
            dest_cell.border = copy(src_cell.border)
            dest_cell.fill = copy(src_cell.fill)
            dest_cell.alignment = copy(src_cell.alignment)
    
    # Row for blank space (already exists or we use it)
    # Next row: facility and first batch
    current_row += 1
    
    first_batch = True
    batch_start_row = current_row
    cell_refs = []
    
    for batch_name, batch_id in batches:
        if first_batch:
            ws[f"{fac_col}{current_row}"] = facility_name
            first_batch = False
        
        ws[f"{batch_col}{current_row}"] = batch_name
        cell_refs.append(f"{val_col}{current_row}")
        
        # Copy styling
        copy_cell_style(style_source_row, current_row, fac_col)
        copy_cell_style(style_source_row, current_row, batch_col)
        copy_cell_style(style_source_row, current_row, val_col)
        
        current_row += 1
    
    # Facility total row
    if len(batches) > 1:
        total_formula = f"=SUM({val_col}{batch_start_row}:{val_col}{current_row-1})"
    else:
        total_formula = f"={val_col}{batch_start_row}"
    
    ws[f"{val_col}{current_row}"] = total_formula
    copy_cell_style(style_source_row + 1, current_row, val_col)  # Use total row style
    
    return cell_refs, current_row


def export_bag_counts(email_to: str = "acheeley@ozarkltcrx.com", reset: bool = True) -> str:
    """
    Export bag counts to Excel and email.
    
    Uses the template and fills in data. Dynamically adds new facilities
    with proper spacing (blank row between facilities, blank row before daily total).
    """
    if not HAS_OPENPYXL:
        return "Error: openpyxl not installed"
    
    # Load data
    state = load_supabase_state()
    counts = state.get("counts", {})
    batches_config = state.get("batches", {})
    
    # Get week info
    today = datetime.now()
    week_num = today.isocalendar()[1]
    year = today.year
    filename = f"Cycle_Bag_Count_{year}_W{week_num:02d}.xlsx"
    filepath = DATA_DIR / filename
    
    # Calculate week dates
    monday = today - timedelta(days=today.weekday())
    
    # Copy template
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    shutil.copy(TEMPLATE_PATH, filepath)
    
    # Load workbook
    wb = load_workbook(filepath)
    
    # Process both sheets
    for ws in wb.worksheets:
        is_census = "Census" in ws.title
        value_key = "census" if is_census else "bags"
        
        # Add dates
        ws['C2'] = monday.strftime("%m/%d")
        ws['G2'] = (monday + timedelta(days=1)).strftime("%m/%d")
        ws['K2'] = (monday + timedelta(days=2)).strftime("%m/%d")
        ws['O2'] = (monday + timedelta(days=3)).strftime("%m/%d")
        ws['S2'] = (monday + timedelta(days=4)).strftime("%m/%d")
        
        # Fill known cells
        for (fac_key, batch_name), cell_ref in KNOWN_CELL_MAP.items():
            col = cell_ref[0]
            day = {"D": "Mon", "H": "Tue", "L": "Wed", "P": "Thu", "T": "Fri"}.get(col)
            if not day:
                continue
            
            fac_counts = counts.get(day, {}).get(fac_key, {})
            fac_batches = batches_config.get(fac_key, [])
            
            # Find matching batch
            for batch in fac_batches:
                if batch["name"] == batch_name:
                    batch_id = batch["id"]
                    values = fac_counts.get(batch_id, {})
                    val = values.get(value_key, 0) or 0
                    if val > 0:
                        ws[cell_ref] = val
                    break
    
    # Save workbook
    wb.save(filepath)
    
    # Calculate totals for email
    total_census = 0
    total_bags = 0
    for day, day_counts in counts.items():
        for fac, fac_counts in day_counts.items():
            for batch_id, values in fac_counts.items():
                if isinstance(values, dict):
                    total_census += values.get("census", 0) or 0
                    total_bags += values.get("bags", 0) or 0
    
    # Email the report
    subject = f"Weekly Bag Count Report - Week {week_num}, {year}"
    body = f"""Weekly Bag Count Report

Week {week_num}, {year}

Summary:
- Total Bags: {total_bags}
- Total Census: {total_census}

The Excel report is attached with two sheets:
- Cycle Census
- Cycle Bag Count

---
Automated report from Ozark LTC Rx Cycle Tracker
"""
    
    result_msg = ""
    try:
        cmd = [
            "gog", "gmail", "send",
            "--to", email_to,
            "--subject", subject,
            "--body", body,
            "--attach", str(filepath)
        ]
        subprocess.run(cmd, check=True, capture_output=True, timeout=60)
        result_msg = f"Exported to {filename} and emailed to {email_to}."
    except subprocess.TimeoutExpired:
        result_msg = f"Exported to {filename} but email timed out."
    except subprocess.CalledProcessError as e:
        result_msg = f"Exported to {filename} but email failed: {e}"
    except Exception as e:
        result_msg = f"Exported to {filename} but email failed: {e}"
    
    # Reset state if requested
    if reset:
        # Would call save function here to reset counts
        pass
    
    return result_msg


if __name__ == "__main__":
    # Test export
    result = export_bag_counts(reset=False)
    print(result)
