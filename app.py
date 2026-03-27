#!/usr/bin/env python3
"""
Ozark LTC Rx Pharmacy Operations Command Center
Run: streamlit run app.py
"""

from __future__ import annotations

import json
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET
from zipfile import ZIP_DEFLATED, ZipFile

import pandas as pd
import streamlit as st
import yaml
import streamlit_authenticator as stauth
from streamlit_autorefresh import st_autorefresh
from dateutil.relativedelta import relativedelta

import supabase_client as supa

try:
    from openpyxl import load_workbook
    from openpyxl import Workbook as _OpenpyxlWorkbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

CYCLE_STAGE_ORDER = [
    "Pulling meds",
    "Exported",
    "Traying",
    "Running in machine",
    "Through machine",
    "Through perl",
    "Bag check completed",
    "Toted",
    "Facility finished",
]

# $$ Tracking stages (subset for cycle fill billing)
CYCLE_DOLLAR_STAGE_ORDER = [
    "Exported",
    "Through machine",
    "Through Perl",
    "Back Checked",
    "Toted",
    "Facility Complete",
]

CYCLE_TEAM_SCHEDULE = {
    "Mon": ["Sunrise Manor", "Oakwood Care", "Pine Valley", "Shady Pines"],
    "Tue": ["Willow Creek", "Maple Ridge", "Cedar Heights", "Harmony House"],
    "Wed": ["Golden Meadows", "River Bend Care", "Silver Oaks", "Evergreen Terrace"],
    "Thu": ["Liberty Place", "Brookside Manor", "Meadow View", "Fox Hollow"],
    "Fri": ["Pleasant Hill", "Autumn Lake", "Grandview Care", "Heritage Pointe"],
}

CYCLE_STAGE_LABELS = {stage: f"{index + 1}. {stage}" for index, stage in enumerate(CYCLE_STAGE_ORDER)}
CYCLE_DOLLAR_STAGE_LABELS = {stage: f"{index + 1}. {stage}" for index, stage in enumerate(CYCLE_DOLLAR_STAGE_ORDER)}

# Frequency options for facilities
FREQUENCY_OPTIONS = ["Weekly", "Every 2 weeks", "Every 4 weeks"]

# Display string for $$ that won't be interpreted as LaTeX (zero-width space between)
DOLLAR_DISPLAY = "$\u200B$"

APP_DIR = Path(__file__).parent
DATA_DIR = APP_DIR / "data"
DATA_FILE = DATA_DIR / "mo_ltc_demo.json"
CYCLE_LOG_FILE = DATA_DIR / "cycle_log.xlsx"

DAY_ABBR_ORDER = ["Mon", "Tue", "Wed", "Thu", "Fri"]
FACILITIES_FILE = DATA_DIR / "facilities.json"
SHARED_STATE_FILE = DATA_DIR / "shared_tracking_state.json"


def load_shared_state() -> dict:
    """Load shared tracking state (Supabase or local JSON fallback)."""
    return supa.load_tracking_state()


def save_shared_state(state: dict) -> None:
    """Save shared tracking state (Supabase or local JSON fallback)."""
    supa.save_tracking_state(state)


def sync_session_with_shared():
    """Sync session state with shared state file."""
    shared = load_shared_state()
    for key in ["cycle_team_tracking", "dollar_tracking", "unlocked_days", "dollar_unlocked_days"]:
        if key in shared and shared[key]:
            st.session_state[key] = shared[key]


def save_session_to_shared():
    """Save relevant session state to shared file."""
    state = {
        "cycle_team_tracking": st.session_state.get("cycle_team_tracking", {}),
        "dollar_tracking": st.session_state.get("dollar_tracking", {}),
        "unlocked_days": st.session_state.get("unlocked_days", []),
        "dollar_unlocked_days": st.session_state.get("dollar_unlocked_days", []),
    }
    save_shared_state(state)


def get_week_dates() -> dict[str, str]:
    """Return dict mapping day abbrev to formatted date string for current week."""
    today = datetime.now()
    # Monday = 0, so we go back to Monday of this week
    monday = today - timedelta(days=today.weekday())
    dates = {}
    for i, day in enumerate(DAY_ABBR_ORDER):
        day_date = monday + timedelta(days=i)
        dates[day] = day_date.strftime("%b %d")
    return dates


def load_facilities_config() -> dict[str, list[dict]]:
    """Load facility schedule from Supabase / JSON, or use default.

    Returns dict mapping day -> list of facility dicts with keys:
    - name: str
    - frequency: str ("Weekly", "Every 2 weeks", "Every 4 weeks")
    """
    # Try Supabase first
    db_data = supa.load_facilities_config_db()
    raw = db_data if db_data else (json.loads(FACILITIES_FILE.read_text()) if FACILITIES_FILE.exists() else None)

    if raw:
        migrated = {}
        for day, facilities in raw.items():
            migrated[day] = []
            for fac in facilities:
                if isinstance(fac, str):
                    migrated[day].append({"name": fac, "frequency": "Weekly"})
                else:
                    if "frequency" not in fac:
                        fac["frequency"] = "Weekly"
                    migrated[day].append(fac)
        return migrated
    # Default schedule
    return {
        day: [{"name": fac, "frequency": "Weekly"} for fac in facilities]
        for day, facilities in CYCLE_TEAM_SCHEDULE.items()
    }


def get_facility_names(config: dict[str, list[dict]]) -> dict[str, list[str]]:
    """Extract just facility names from config (for backward compat)."""
    return {day: [f["name"] for f in facs] for day, facs in config.items()}


def save_facilities_config(config: dict[str, list[dict]]) -> None:
    """Save facility schedule to Supabase and local JSON."""
    supa.save_facilities_config_db(config)
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    FACILITIES_FILE.write_text(json.dumps(config, indent=2))


def get_current_week_number() -> int:
    """Get ISO week number for current date."""
    return datetime.now().isocalendar()[1]


def get_week_start_date(target_date: datetime | None = None) -> datetime:
    """Get the Monday of the week containing target_date (or today)."""
    if target_date is None:
        target_date = datetime.now()
    return target_date - timedelta(days=target_date.weekday())


def facility_active_this_week(frequency: str, start_date: str | None = None) -> bool:
    """Check if facility is active this week based on frequency and start date.
    
    Args:
        frequency: "Weekly", "Every 2 weeks", or "Every 4 weeks"
        start_date: ISO date string (YYYY-MM-DD) of first run, or None for Weekly
    """
    if frequency == "Weekly":
        return True
    
    if not start_date:
        # No start date set, default to active (treat as Weekly)
        return True
    
    try:
        start = datetime.strptime(start_date, "%Y-%m-%d")
    except (ValueError, TypeError):
        return True
    
    # Get Monday of start week and current week
    start_week_monday = get_week_start_date(start)
    current_week_monday = get_week_start_date()
    
    # Calculate weeks since start
    weeks_diff = (current_week_monday - start_week_monday).days // 7
    
    if frequency == "Every 2 weeks":
        return weeks_diff % 2 == 0
    elif frequency == "Every 4 weeks":
        return weeks_diff % 4 == 0
    
    return True


def get_next_run_date(frequency: str, start_date: str | None, day_abbr: str) -> str:
    """Get the next run date for a facility based on frequency."""
    if frequency == "Weekly":
        # Next occurrence of this day
        today = datetime.now()
        day_idx = DAY_ABBR_ORDER.index(day_abbr)
        days_until = (day_idx - today.weekday()) % 7
        if days_until == 0 and today.hour >= 18:  # Past 6pm, show next week
            days_until = 7
        next_date = today + timedelta(days=days_until)
        return next_date.strftime("%b %d")
    
    if not start_date:
        return "Not set"
    
    try:
        start = datetime.strptime(start_date, "%Y-%m-%d")
    except (ValueError, TypeError):
        return "Invalid"
    
    today = datetime.now()
    interval = 14 if frequency == "Every 2 weeks" else 28
    
    # Find next occurrence on or after today
    current = start
    while current < today:
        current += timedelta(days=interval)
    
    return current.strftime("%b %d")


def _today_abbr() -> str:
    return datetime.now().strftime("%a")[:3]


def get_visible_days(tracking_state: dict) -> list[str]:
    """Return days to show: today + any prior day with incomplete facilities."""
    today = _today_abbr()
    if today not in DAY_ABBR_ORDER:
        return DAY_ABBR_ORDER  # weekend: show all
    today_idx = DAY_ABBR_ORDER.index(today)
    visible = []
    for i, day in enumerate(DAY_ABBR_ORDER):
        if i == today_idx:
            visible.append(day)
            continue
        if i > today_idx:
            continue  # future days hidden
        # Past day: show only if has incomplete facilities
        day_state = tracking_state.get(day, {})
        for fac in CYCLE_TEAM_SCHEDULE.get(day, []):
            c, t = stage_counts(day_state.get(fac, {}))
            if c < t:
                visible.append(day)
                break
    return visible


def _excel_col_name(index: int) -> str:
    result = ""
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _xlsx_escape(value: Any) -> str:
    text = "" if value is None else str(value)
    return (
        text.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def _build_sheet_xml(rows: list[list[Any]]) -> str:
    dimension = f"A1:E{max(len(rows), 1)}"
    row_xml = []
    for row_idx, row in enumerate(rows, start=1):
        cell_xml = []
        for col_idx, value in enumerate(row, start=1):
            cell_ref = f"{_excel_col_name(col_idx)}{row_idx}"
            cell_xml.append(
                f'<c r="{cell_ref}" t="inlineStr"><is><t>{_xlsx_escape(value)}</t></is></c>'
            )
        row_xml.append(f'<row r="{row_idx}">{"".join(cell_xml)}</row>')
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        f'<dimension ref="{dimension}"/>'
        '<sheetViews><sheetView workbookViewId="0"/></sheetViews>'
        '<sheetFormatPr defaultRowHeight="15"/>'
        f'<sheetData>{"".join(row_xml)}</sheetData>'
        '<pageMargins left="0.7" right="0.7" top="0.75" bottom="0.75" header="0.3" footer="0.3"/>'
        '</worksheet>'
    )


def _write_basic_xlsx(path: Path, rows: list[list[Any]]) -> None:
    content_types = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
  <Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
  <Override PartName="/docProps/core.xml" ContentType="application/vnd.openxmlformats-package.core-properties+xml"/>
  <Override PartName="/docProps/app.xml" ContentType="application/vnd.openxmlformats-officedocument.extended-properties+xml"/>
</Types>'''
    rels = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
  <Relationship Id="rId2" Type="http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties" Target="docProps/core.xml"/>
  <Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/extended-properties" Target="docProps/app.xml"/>
</Relationships>'''
    workbook = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <sheets><sheet name="Cycle Log" sheetId="1" r:id="rId1"/></sheets>
</workbook>'''
    workbook_rels = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>
</Relationships>'''
    core = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties" xmlns:dc="http://purl.org/dc/elements/1.1/" xmlns:dcterms="http://purl.org/dc/terms/" xmlns:dcmitype="http://purl.org/dc/dcmitype/" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
  <dc:title>Cycle Log</dc:title>
  <dc:creator>Command Center</dc:creator>
</cp:coreProperties>'''
    app = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/extended-properties" xmlns:vt="http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes">
  <Application>Command Center</Application>
</Properties>'''

    with ZipFile(path, "w", compression=ZIP_DEFLATED) as zf:
        zf.writestr("[Content_Types].xml", content_types)
        zf.writestr("_rels/.rels", rels)
        zf.writestr("docProps/core.xml", core)
        zf.writestr("docProps/app.xml", app)
        zf.writestr("xl/workbook.xml", workbook)
        zf.writestr("xl/_rels/workbook.xml.rels", workbook_rels)
        zf.writestr("xl/worksheets/sheet1.xml", _build_sheet_xml(rows))


def _read_basic_xlsx_rows(path: Path) -> list[list[str]]:
    ns = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    with ZipFile(path) as zf:
        sheet_xml = zf.read("xl/worksheets/sheet1.xml")
    root = ET.fromstring(sheet_xml)
    rows: list[list[str]] = []
    for row in root.findall("main:sheetData/main:row", ns):
        current: list[str] = []
        for cell in row.findall("main:c", ns):
            inline = cell.find("main:is/main:t", ns)
            value = inline.text if inline is not None and inline.text is not None else ""
            current.append(value)
        rows.append(current)
    return rows


def ensure_cycle_log_file() -> None:
    """Create the Excel audit file with headers if it does not already exist."""
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    headers = ["Facility", "Stage", "Initials", "Date", "Time"]
    if CYCLE_LOG_FILE.exists():
        return
    if HAS_OPENPYXL:
        wb = _OpenpyxlWorkbook()
        ws = wb.active
        ws.title = "Cycle Log"
        ws.append(headers)
        wb.save(CYCLE_LOG_FILE)
        return
    _write_basic_xlsx(CYCLE_LOG_FILE, [headers])


def log_stage_to_excel(facility: str, stage: str, initials: str) -> None:
    """Append a completion record to audit log (Supabase + Excel)."""
    supa.log_audit_entry(facility, stage, initials)
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    now = datetime.now()
    row = [facility, stage, initials, now.strftime("%Y-%m-%d"), now.strftime("%H:%M:%S")]
    headers = ["Facility", "Stage", "Initials", "Date", "Time"]
    if HAS_OPENPYXL:
        if CYCLE_LOG_FILE.exists():
            wb = load_workbook(CYCLE_LOG_FILE)
            ws = wb.active
        else:
            wb = _OpenpyxlWorkbook()
            ws = wb.active
            ws.title = "Cycle Log"
            ws.append(headers)
        ws.append(row)
        wb.save(CYCLE_LOG_FILE)
        return

    existing_rows = _read_basic_xlsx_rows(CYCLE_LOG_FILE) if CYCLE_LOG_FILE.exists() else []
    if not existing_rows:
        existing_rows = [headers]
    existing_rows.append(row)
    _write_basic_xlsx(CYCLE_LOG_FILE, existing_rows)

CUSTOM_CSS = """
<style>
    .stApp {
        background: linear-gradient(180deg, #f8fbff 0%, #eef4ff 100%);
    }
    section[data-testid="stSidebar"] {
        background: #0f172a !important;
    }
    section[data-testid="stSidebar"] * {
        color: white !important;
    }
    /* Sidebar nav button styling */
    section[data-testid="stSidebar"] button[kind="secondary"] {
        background: #2563eb !important;
        color: white !important;
        border: none !important;
        margin-bottom: 4px !important;
    }
    section[data-testid="stSidebar"] button[kind="secondary"]:hover {
        background: #1d4ed8 !important;
        color: white !important;
    }
    section[data-testid="stSidebar"] button[kind="primary"] {
        background: #1e40af !important;
        color: white !important;
        border-left: 4px solid white !important;
        margin-bottom: 4px !important;
    }
    section[data-testid="stSidebar"] button[kind="primary"]:hover {
        background: #1e3a8a !important;
    }

    .hero-card, .block-card {
        background: white;
        border: 1px solid #dbe4f0;
        border-radius: 18px;
        padding: 18px 20px;
        box-shadow: 0 6px 18px rgba(15, 23, 42, 0.06);
        margin-bottom: 12px;
    }
    .hero-title {
        font-size: 1.6rem;
        font-weight: 800;
        color: #0f172a;
    }
    .muted {
        color: #64748b;
        font-size: 0.95rem;
    }
    .pill {
        display: inline-block;
        padding: 6px 10px;
        border-radius: 999px;
        font-size: 12px;
        font-weight: 700;
        margin-right: 8px;
        margin-bottom: 8px;
        background: #e0e7ff;
        color: #3730a3;
    }
</style>
"""

RISK_ORDER = {"High": 0, "Medium": 1, "Low": 2}
STATUS_COLORS = {
    "High": "#dc2626",
    "Medium": "#d97706",
    "Low": "#059669",
    "At Risk": "#dc2626",
    "Watch": "#d97706",
    "On Track": "#059669",
    "Delayed": "#dc2626",
    "In progress": "#2563eb",
    "Complete": "#059669",
}


@st.cache_data
def load_demo_data() -> dict[str, Any]:
    return json.loads(DATA_FILE.read_text())


def risk_badge(label: str) -> str:
    color = STATUS_COLORS.get(label, "#475569")
    return (
        f"<span style='display:inline-block;padding:4px 10px;border-radius:999px;"
        f"background:{color};color:white;font-size:12px;font-weight:700;'>{label}</span>"
    )


def metric_block(label: str, value: Any, help_text: str) -> None:
    st.metric(label, value, help=help_text)


def build_cycle_team_state() -> dict[str, dict[str, dict[str, str]]]:
    return {
        day: {
            facility: {stage: "" for stage in CYCLE_STAGE_ORDER}
            for facility in facilities
        }
        for day, facilities in CYCLE_TEAM_SCHEDULE.items()
    }


def stage_counts(stage_map: dict[str, str], stage_order: list[str] = None) -> tuple[int, int]:
    if stage_order is None:
        stage_order = CYCLE_STAGE_ORDER
    completed = sum(1 for stage in stage_order if stage_map.get(stage, ""))
    return completed, len(stage_order)


def cycle_status_label(stage_map: dict[str, str], stage_order: list[str] = None) -> str:
    if stage_order is None:
        stage_order = CYCLE_STAGE_ORDER
    completed, total = stage_counts(stage_map, stage_order)
    if completed == 0:
        return "Not Started"
    if completed >= total:
        return "Completed"

    for stage in reversed(stage_order):
        if stage_map.get(stage, ""):
            return stage

    return "Not Started"


def cycle_status_color(status: str) -> str:
    return {
        "Not Started": "#94a3b8",
        "Completed": "#059669",
    }.get(status, "#2563eb")


def render_cycle_facility(day: str, facility: str, stage_map: dict[str, str]) -> None:
    completed, total = stage_counts(stage_map)
    progress_pct = completed / total
    status = cycle_status_label(stage_map)
    accent = cycle_status_color(status)

    # Build expander label with progress info
    pct_display = int(progress_pct * 100)
    expander_label = f"**{facility}** — {status} ({pct_display}%)"
    
    with st.expander(expander_label, expanded=False):
        initials_summary = ", ".join(
            f"{stage}: {initials}"
            for stage, initials in stage_map.items()
            if initials
        )
        
        progress_col, summary_col = st.columns([1.25, 1])
        with progress_col:
            st.progress(progress_pct, text=f"{pct_display}% complete")
        with summary_col:
            if status == "Completed":
                st.success("Completed and ready to roll.")
            elif status == "Not Started":
                st.caption("No stages marked yet.")
            else:
                st.info(f"Latest: {status}")
            if initials_summary:
                st.caption(f"By: {initials_summary}")

        stage_cols = st.columns(3)
        for idx, stage in enumerate(CYCLE_STAGE_ORDER):
            initials_key = f"cycle_initials::{day}::{facility}::{stage}"
            form_key = f"cycle_form::{day}::{facility}::{stage}"
            completed_initials = stage_map.get(stage, "")
            with stage_cols[idx % 3]:
                if completed_initials:
                    st.success(f"{CYCLE_STAGE_LABELS[stage]}\n\n✓ {completed_initials}")
                else:
                    with st.form(form_key, clear_on_submit=True):
                        initials_value = st.text_input(
                            CYCLE_STAGE_LABELS[stage],
                            max_chars=4,
                            key=initials_key,
                            placeholder="AB",
                        )
                        submitted = st.form_submit_button("Mark complete", use_container_width=True)
                        if submitted:
                            cleaned = initials_value.strip().upper()
                            if cleaned:
                                stage_map[stage] = cleaned
                                st.session_state.cycle_team_tracking[day][facility][stage] = cleaned
                                log_stage_to_excel(facility, stage, cleaned)
                                save_session_to_shared()  # Sync to shared state
                                st.rerun()
                            else:
                                st.warning("Enter initials to complete the stage.")


def render_dollar_facility(day: str, facility: str, stage_map: dict[str, str]) -> None:
    """Render a facility card for $$ tracking with different stages."""
    # Check for "No meds" bypass
    no_meds = stage_map.get("_no_meds", False)
    
    if no_meds:
        # Show as bypassed
        expander_label = f"**{facility} {DOLLAR_DISPLAY}** — No Meds ⏭️"
        with st.expander(expander_label, expanded=False):
            st.info("🚫 No meds to run — bypassed")
            if st.button(f"↩️ Undo bypass", key=f"undo_nomeds_{day}_{facility}"):
                del st.session_state.dollar_tracking[day][facility]["_no_meds"]
                save_session_to_shared()
                st.rerun()
        return
    
    completed, total = stage_counts(stage_map, CYCLE_DOLLAR_STAGE_ORDER)
    progress_pct = completed / total if total > 0 else 0
    status = cycle_status_label(stage_map, CYCLE_DOLLAR_STAGE_ORDER)
    
    # Build expander label with $$ suffix
    pct_display = int(progress_pct * 100)
    expander_label = f"**{facility} {DOLLAR_DISPLAY}** — {status} ({pct_display}%)"
    
    with st.expander(expander_label, expanded=False):
        # No meds bypass button at top
        bypass_col, spacer_col = st.columns([1, 2])
        with bypass_col:
            if st.button("🚫 No meds to run", key=f"nomeds_{day}_{facility}", use_container_width=True):
                st.session_state.dollar_tracking[day][facility]["_no_meds"] = True
                log_stage_to_excel(f"{facility} $$", "NO MEDS", "BYPASS")
                save_session_to_shared()
                st.rerun()
        
        initials_summary = ", ".join(
            f"{stage}: {initials}"
            for stage, initials in stage_map.items()
            if initials and not stage.startswith("_")
        )
        
        progress_col, summary_col = st.columns([1.25, 1])
        with progress_col:
            st.progress(progress_pct, text=f"{pct_display}% complete")
        with summary_col:
            if status == "Completed":
                st.success(f"{DOLLAR_DISPLAY} Complete!")
            elif status == "Not Started":
                st.caption("No stages marked yet.")
            else:
                st.info(f"Latest: {status}")
            if initials_summary:
                st.caption(f"By: {initials_summary}")

        stage_cols = st.columns(3)
        for idx, stage in enumerate(CYCLE_DOLLAR_STAGE_ORDER):
            initials_key = f"dollar_initials::{day}::{facility}::{stage}"
            form_key = f"dollar_form::{day}::{facility}::{stage}"
            completed_initials = stage_map.get(stage, "")
            with stage_cols[idx % 3]:
                if completed_initials:
                    st.success(f"{CYCLE_DOLLAR_STAGE_LABELS[stage]}\n\n✓ {completed_initials}")
                else:
                    with st.form(form_key, clear_on_submit=True):
                        initials_value = st.text_input(
                            CYCLE_DOLLAR_STAGE_LABELS[stage],
                            max_chars=4,
                            key=initials_key,
                            placeholder="AB",
                        )
                        submitted = st.form_submit_button("Mark complete", use_container_width=True)
                        if submitted:
                            cleaned = initials_value.strip().upper()
                            if cleaned:
                                stage_map[stage] = cleaned
                                st.session_state.dollar_tracking[day][facility][stage] = cleaned
                                log_stage_to_excel(f"{facility} $$", stage, cleaned)
                                save_session_to_shared()
                                st.rerun()
                            else:
                                st.warning("Enter initials to complete the stage.")


ensure_cycle_log_file()

st.set_page_config(page_title="Ozark LTC Rx Ops Center", page_icon="💊", layout="wide")

# --- Authentication (with Supabase persistence for 30-day sessions) ---
CONFIG_PATH = APP_DIR / "config.yaml"


def load_merged_config() -> dict:
    """Load config from Supabase (if available), merged with local config.yaml.
    
    Supabase config takes precedence for credentials/permissions, ensuring
    user accounts persist across Streamlit Cloud restarts.
    """
    # Start with local config.yaml as base
    local_config = {}
    if CONFIG_PATH.exists():
        with open(CONFIG_PATH) as f:
            local_config = yaml.safe_load(f) or {}
    
    # Try to load from Supabase
    db_config = supa.load_auth_config_db()
    
    if db_config:
        # Supabase config takes precedence for credentials and permissions
        merged = local_config.copy()
        if "credentials" in db_config:
            merged["credentials"] = db_config["credentials"]
        if "permissions" in db_config:
            merged["permissions"] = db_config["permissions"]
        return merged
    
    return local_config


def save_config(config: dict) -> None:
    """Save config to both Supabase (for persistence) and local file.
    
    This ensures user accounts survive Streamlit Cloud restarts.
    """
    # Save to Supabase first (this persists across restarts)
    supa.save_auth_config_db(config)
    
    # Also save locally (for immediate use and backup)
    with open(CONFIG_PATH, "w") as f:
        yaml.dump(config, f, default_flow_style=False)


# Load merged config (Supabase + local)
config = load_merged_config()

if config and config.get('credentials') and config.get('cookie'):
    authenticator = stauth.Authenticate(
        config['credentials'],
        config['cookie']['name'],
        config['cookie']['key'],
        config['cookie']['expiry_days'],  # 30 days as configured
    )
    
    authenticator.login(location='main')
    
    if st.session_state.get("authentication_status") is None:
        st.warning("Please enter your username and password")
        st.stop()
    elif st.session_state.get("authentication_status") is False:
        st.error("Username/password is incorrect")
        st.stop()
    
    # User is authenticated - show logout in sidebar later
else:
    st.error("No authentication config found. Please check config.yaml and Supabase.")
    st.stop()
# --- End Authentication ---

# --- Auto-refresh for real-time sync (every 30 seconds) ---
st_autorefresh(interval=30000, limit=None, key="auto_refresh")

# --- Sync with shared state ---
sync_session_with_shared()

st.markdown(CUSTOM_CSS, unsafe_allow_html=True)

data = load_demo_data()
facilities = pd.DataFrame(data["facilities"])
daily_ops = pd.DataFrame(data["daily_ops"])
adt = pd.DataFrame(data["adt_events"])
cycle_fill = pd.DataFrame(data["cycle_fill"])
delivery = pd.DataFrame(data["delivery_runs"])
emergency_kits = pd.DataFrame(data["emergency_kits"])

facility_names = ["All facilities", *facilities["facility"].tolist()]
selected_facility = "All facilities"

# --- Permission System ---
ALL_PAGES = [
    "Dashboard",
    "Cycle Team",
    "Facility Management",
    "Pharmacy Management",
    "QA",
    "Data Explorer",
    "User Management",
]

def get_user_permissions():
    """Get list of pages the current user can access."""
    if not CONFIG_PATH.exists():
        return ALL_PAGES  # No auth = full access
    
    username = st.session_state.get("username")
    if not username:
        return []
    
    user_data = config.get("credentials", {}).get("usernames", {}).get(username, {})
    role = user_data.get("role", "user")
    
    permissions = config.get("permissions", {})
    
    # Check user-specific permissions first
    user_perms = permissions.get("users", {}).get(username)
    if user_perms:
        return user_perms
    
    # Fall back to role permissions
    role_perms = permissions.get("roles", {}).get(role, [])
    if "all" in role_perms:
        return ALL_PAGES
    return role_perms

def is_admin_user():
    if not CONFIG_PATH.exists():
        return True
    username = st.session_state.get("username")
    if username and username in config.get("credentials", {}).get("usernames", {}):
        return config["credentials"]["usernames"][username].get("role") == "admin"
    return False

# Get pages this user can see
allowed_pages = get_user_permissions()

with st.sidebar:
    st.title("💊 Ozark LTC Rx")
    
    # Show logged-in user and logout button
    if CONFIG_PATH.exists() and st.session_state.get("authentication_status"):
        st.caption(f"👤 **{st.session_state.get('name', 'User')}**")
        # Red logout button using custom HTML
        st.markdown("""
            <style>
            div[data-testid="stSidebar"] > div > div > div > div:nth-child(3) button {
                background: #dc2626 !important;
                color: white !important;
            }
            div[data-testid="stSidebar"] > div > div > div > div:nth-child(3) button:hover {
                background: #b91c1c !important;
            }
            </style>
        """, unsafe_allow_html=True)
        authenticator.logout("🚪 Logout", "sidebar")
        st.divider()
    
    # Navigation
    st.subheader("Navigation")
    
    # Only show pages user has access to
    if "current_page" not in st.session_state:
        st.session_state.current_page = allowed_pages[0] if allowed_pages else "Dashboard"
    
    # Ensure current page is allowed
    if st.session_state.current_page not in allowed_pages and allowed_pages:
        st.session_state.current_page = allowed_pages[0]
    
    for page in ALL_PAGES:
        if page in allowed_pages:
            if st.button(
                f"{'→ ' if st.session_state.current_page == page else '   '}{page}",
                key=f"nav_{page}",
                use_container_width=True,
                type="primary" if st.session_state.current_page == page else "secondary"
            ):
                st.session_state.current_page = page
                st.rerun()
    
    st.divider()
    st.caption("Concrete pharmacy operations prototype")
    selected_facility = st.selectbox("Facility filter", facility_names)
    st.divider()
    # Load master facility count for sidebar
    _sidebar_master_facs = supa.load_master_facilities()
    st.caption(f"Facilities: {len(_sidebar_master_facs)} · ADT: {len(adt)}")
    st.caption(f"Routes: {len(delivery)} · E-Kits: {len(emergency_kits)}")

if selected_facility != "All facilities":
    facilities_view = facilities[facilities["facility"] == selected_facility].copy()
    adt_view = adt[adt["facility"] == selected_facility].copy()
    cycle_view = cycle_fill[cycle_fill["facility"] == selected_facility].copy()
    ekit_view = emergency_kits[emergency_kits["facility"] == selected_facility].copy()
else:
    facilities_view = facilities.copy()
    adt_view = adt.copy()
    cycle_view = cycle_fill.copy()
    ekit_view = emergency_kits.copy()

high_risk_count = int((facilities_view["risk_level"] == "High").sum())
pending_orders = int(facilities_view["pending_orders"].sum())
overdue_first_doses = int(facilities_view["overdue_first_doses"].sum())
adt_admissions = int((adt_view["event_type"] == "Admission").sum())
packed_total = int(cycle_view["packed"].sum()) if not cycle_view.empty else 0
residents_due_total = int(cycle_view["residents_due"].sum()) if not cycle_view.empty else 0
fill_completion = round((packed_total / residents_due_total) * 100) if residents_due_total else 0

# Get current page
current_page = st.session_state.get("current_page", "Dashboard")

# --- PAGE: Dashboard ---
if current_page == "Dashboard":
    st.title("Dashboard")
    st.caption(f"Overview as of {datetime.now().strftime('%A %Y-%m-%d %H:%M')}")

    # Load facility config and tracking state
    _dash_facilities_raw = load_facilities_config()
    _dash_facilities: dict[str, list[str]] = {}
    for _d, _fl in _dash_facilities_raw.items():
        _dash_facilities[_d] = [
            f["name"] for f in _fl
            if facility_active_this_week(f.get("frequency", "Weekly"), f.get("start_date"))
        ]

    _dash_tracking = st.session_state.get("cycle_team_tracking", {})

    today_key = _today_abbr()
    today_idx = DAY_ABBR_ORDER.index(today_key) if today_key in DAY_ABBR_ORDER else -1
    now_hour = datetime.now().hour + datetime.now().minute / 60.0

    def get_status_with_timing(fac: str, stage_map: dict, is_overdue: bool = False) -> str:
        """Determine status based on 30-min window vs 4-week average.
        
        Returns: Not Started, Running Behind, On Time, Ahead of Schedule, or Completed
        """
        c, t = stage_counts(stage_map)
        
        if c == 0:
            return "Not Started"
        if c >= t:
            return "Completed"
        
        # In progress - check timing
        if is_overdue:
            return "Running Behind"
        
        avg_hour = supa.get_average_completion_hour(fac, "Facility finished", weeks=4)
        if avg_hour is None:
            return "On Time"  # No historical data
        
        diff_minutes = (now_hour - avg_hour) * 60
        
        if diff_minutes >= 30:
            return "Running Behind"
        elif diff_minutes <= -30:
            return "Ahead of Schedule"
        else:
            return "On Time"

    def get_current_task(stage_map: dict) -> str:
        """Get the current task (first incomplete stage) or 'Done'."""
        for stage in CYCLE_STAGE_ORDER:
            if not stage_map.get(stage, ""):
                return stage
        return "Done"

    # Collect rows: today's facilities + previous-day incomplete
    dash_rows: list[dict] = []
    seen: set[tuple[str, str]] = set()

    # Previous days incomplete (only show last 2 days, hide older)
    if today_key in DAY_ABBR_ORDER:
        recent_cutoff = max(0, today_idx - 2)
        for prior_day in DAY_ABBR_ORDER[recent_cutoff:today_idx]:
            for fac in _dash_facilities.get(prior_day, []):
                sm = _dash_tracking.get(prior_day, {}).get(fac, {})
                c, t = stage_counts(sm)
                if c < t:
                    status = get_status_with_timing(fac, sm, is_overdue=True)
                    current_task = get_current_task(sm)
                    dash_rows.append({
                        "Facility": f"{fac} ({prior_day})",
                        "Tasks Completed": f"{c}/{t}",
                        "Current Task": current_task,
                        "Status": status,
                        "_status": status,
                    })
                    seen.add((prior_day, fac))

    # Today's facilities
    for fac in _dash_facilities.get(today_key, []):
        if (today_key, fac) in seen:
            continue
        sm = _dash_tracking.get(today_key, {}).get(fac, {})
        c, t = stage_counts(sm)
        status = get_status_with_timing(fac, sm, is_overdue=False)
        current_task = get_current_task(sm)
        
        dash_rows.append({
            "Facility": fac,
            "Tasks Completed": f"{c}/{t}",
            "Current Task": current_task,
            "Status": status,
            "_status": status,
        })

    if dash_rows:
        # Summary metrics
        total = len(dash_rows)
        completed = sum(1 for r in dash_rows if r["_status"] == "Completed")
        behind = sum(1 for r in dash_rows if r["_status"] == "Running Behind")
        ahead = sum(1 for r in dash_rows if r["_status"] == "Ahead of Schedule")
        
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Total Facilities", total)
        m2.metric("Completed", completed)
        m3.metric("Running Behind", behind)
        m4.metric("Ahead of Schedule", ahead)

        # Status colors for row highlighting
        STATUS_COLORS = {
            "Not Started": {"bg": "#f8fafc", "border": "#e2e8f0", "text": "#64748b"},
            "Running Behind": {"bg": "#fef2f2", "border": "#fca5a5", "text": "#dc2626"},
            "On Time": {"bg": "#ffffff", "border": "#dbe4f0", "text": "#1e40af"},
            "Ahead of Schedule": {"bg": "#f0fdf4", "border": "#86efac", "text": "#059669"},
            "Completed": {"bg": "#f0fdf4", "border": "#86efac", "text": "#059669"},
        }

        # Table header
        st.markdown("""
        <style>
        .dash-header {
            display: flex;
            font-weight: 700;
            padding: 12px 16px;
            background: #f1f5f9;
            border-radius: 8px 8px 0 0;
            border: 1px solid #e2e8f0;
            margin-top: 16px;
        }
        .dash-header > div { flex: 1; }
        .dash-row {
            display: flex;
            padding: 12px 16px;
            border-left: 1px solid;
            border-right: 1px solid;
            border-bottom: 1px solid;
        }
        .dash-row:last-child { border-radius: 0 0 8px 8px; }
        .dash-row > div { flex: 1; }
        </style>
        """, unsafe_allow_html=True)
        
        # Header row
        st.markdown("""
        <div class="dash-header">
            <div>Facility</div>
            <div>Tasks Completed</div>
            <div>Current Task</div>
            <div>Status</div>
        </div>
        """, unsafe_allow_html=True)
        
        # Data rows with color coding
        for row in dash_rows:
            colors = STATUS_COLORS.get(row["_status"], STATUS_COLORS["On Time"])
            st.markdown(f"""
            <div class="dash-row" style="background:{colors['bg']};border-color:{colors['border']};">
                <div style="font-weight:600;">{row['Facility']}</div>
                <div>{row['Tasks Completed']}</div>
                <div>{row['Current Task']}</div>
                <div style="color:{colors['text']};font-weight:600;">{row['Status']}</div>
            </div>
            """, unsafe_allow_html=True)
    else:
        st.info("No facilities scheduled for today.")

# --- PAGE: Cycle Team (combined with internal tabs) ---
if current_page == "Cycle Team":
    st.title("Cycle Team")
    
    # Load facility config once for all tabs
    _cycle_facilities_raw = load_facilities_config()
    _cycle_facilities = {}
    for _day, _fac_list in _cycle_facilities_raw.items():
        _cycle_facilities[_day] = [
            f["name"] for f in _fac_list 
            if facility_active_this_week(f.get("frequency", "Weekly"), f.get("start_date"))
        ]
    
    # Initialize tracking states
    if "cycle_team_tracking" not in st.session_state:
        st.session_state.cycle_team_tracking = {
            day: {fac: {stage: "" for stage in CYCLE_STAGE_ORDER} for fac in facs}
            for day, facs in _cycle_facilities.items()
        }
    if "dollar_tracking" not in st.session_state:
        st.session_state.dollar_tracking = {
            day: {fac: {stage: "" for stage in CYCLE_DOLLAR_STAGE_ORDER} for fac in facs}
            for day, facs in _cycle_facilities.items()
        }
    if "unlocked_days" not in st.session_state:
        st.session_state.unlocked_days = []
    if "dollar_unlocked_days" not in st.session_state:
        st.session_state.dollar_unlocked_days = []
    
    # Create internal tabs - Dashboard first
    cycle_tab1, cycle_tab2, cycle_tab3, cycle_tab4 = st.tabs([
        "📊 Cycle Team Dashboard",
        "📋 Cycle Team Tracking",
        "💰 Cycle High Dollar Tracking",
        "📦 Cycle Bag Count"
    ])
    
    # ============ TAB 1: DASHBOARD ============
    with cycle_tab1:
        week_num = get_current_week_number()
        tracking = st.session_state.cycle_team_tracking
        dollar_tracking = st.session_state.dollar_tracking
        
        st.markdown("### Cycle Team Dashboard")
        st.caption(f"Supervisor summary as of {datetime.now().strftime('%A %Y-%m-%d %H:%M')} (Week {week_num})")
        
        today_key = _today_abbr()
        if today_key in _cycle_facilities and _cycle_facilities[today_key]:
            st.markdown(f"#### Today ({today_key})")
            today_rows = []
            for fac in _cycle_facilities[today_key]:
                stage_map = tracking.get(today_key, {}).get(fac, {})
                c, t = stage_counts(stage_map)
                status = cycle_status_label(stage_map)
                pct = int(c / t * 100) if t else 0
                
                dollar_map = dollar_tracking.get(today_key, {}).get(fac, {})
                if dollar_map.get("_no_meds"):
                    dollar_status = "No Meds"
                else:
                    dc, dt = stage_counts(dollar_map, CYCLE_DOLLAR_STAGE_ORDER)
                    if dc == 0:
                        dollar_status = "Not Started"
                    elif dc >= dt:
                        dollar_status = "Completed"
                    else:
                        dollar_status = "In Progress"
                
                today_rows.append({
                    "Facility": fac,
                    "Progress": f"{c}/{t} ({pct}%)",
                    "Status": status,
                    "High Dollar": dollar_status
                })
            
            today_df = pd.DataFrame(today_rows)
            done_count = sum(1 for r in today_rows if r["Status"] == "Completed")
            dollar_done = sum(1 for r in today_rows if r["High Dollar"] in ["Completed", "No Meds"])
            
            m1, m2, m3, m4 = st.columns(4)
            m1.metric("Total Facilities", len(today_rows))
            m2.metric("Cycle Complete", done_count)
            m3.metric("High Dollar Complete", dollar_done)
            m4.metric("Remaining", len(today_rows) - done_count)
            st.dataframe(today_df, use_container_width=True, hide_index=True)
        else:
            st.info("No cycle schedule for today.")
        
        # Overdue section (only show last 2 days, hide older)
        overdue_rows = []
        if today_key in DAY_ABBR_ORDER:
            today_idx = DAY_ABBR_ORDER.index(today_key)
            recent_cutoff = max(0, today_idx - 2)
            for day in DAY_ABBR_ORDER[recent_cutoff:today_idx]:
                day_state = tracking.get(day, {})
                for fac in _cycle_facilities.get(day, []):
                    c, t = stage_counts(day_state.get(fac, {}))
                    if c < t:
                        overdue_rows.append({"Day": day, "Facility": fac, "Progress": f"{c}/{t}"})
        
        if overdue_rows:
            st.markdown("#### Overdue / Carryover")
            st.warning(f"{len(overdue_rows)} facility(ies) from prior days still incomplete.")
            st.dataframe(pd.DataFrame(overdue_rows), use_container_width=True, hide_index=True)
        else:
            st.success("No overdue facilities from prior days.")
    
    # ============ TAB 2: CYCLE TEAM TRACKING ============
    with cycle_tab2:
        cycle_facilities = _cycle_facilities
        
        for day, facs in cycle_facilities.items():
            if day not in st.session_state.cycle_team_tracking:
                st.session_state.cycle_team_tracking[day] = {}
            for fac in facs:
                if fac not in st.session_state.cycle_team_tracking[day]:
                    st.session_state.cycle_team_tracking[day][fac] = {stage: "" for stage in CYCLE_STAGE_ORDER}
        
        st.markdown("### Cycle Team Tracking")
        today_abbr = _today_abbr()
        week_dates = get_week_dates()
        today_idx = DAY_ABBR_ORDER.index(today_abbr) if today_abbr in DAY_ABBR_ORDER else -1
        
        def get_visible_days_cycle(ts, fc):
            if today_abbr not in DAY_ABBR_ORDER:
                return DAY_ABBR_ORDER
            visible = []
            for i, d in enumerate(DAY_ABBR_ORDER):
                if i == today_idx:
                    visible.append(d)
                elif i < today_idx:
                    for f in fc.get(d, []):
                        c, t = stage_counts(ts.get(d, {}).get(f, {}))
                        if c < t:
                            visible.append(d)
                            break
            return visible
        
        visible_days = get_visible_days_cycle(st.session_state.cycle_team_tracking, cycle_facilities)
        if not visible_days:
            visible_days = [today_abbr] if today_abbr in DAY_ABBR_ORDER else DAY_ABBR_ORDER
        
        for unlocked in st.session_state.unlocked_days:
            if unlocked not in visible_days and unlocked in DAY_ABBR_ORDER:
                visible_days.append(unlocked)
        visible_days = sorted(set(visible_days), key=lambda d: DAY_ABBR_ORDER.index(d) if d in DAY_ABBR_ORDER else 99)
        
        future_days = DAY_ABBR_ORDER[today_idx + 1:] if today_idx >= 0 else []
        next_unlockable = next((fd for fd in future_days if fd not in st.session_state.unlocked_days), None)
        
        st.caption("Click a facility to expand and mark stages complete.")
        btn_cols = st.columns(5)
        col_idx = 0
        for unlocked in st.session_state.unlocked_days:
            if col_idx < 5:
                with btn_cols[col_idx]:
                    if st.button(f"🔒 Lock {unlocked}", key=f"relock_{unlocked}", use_container_width=True):
                        st.session_state.unlocked_days.remove(unlocked)
                        st.session_state.unlocked_days = [d for d in st.session_state.unlocked_days if DAY_ABBR_ORDER.index(d) < DAY_ABBR_ORDER.index(unlocked)]
                        save_session_to_shared()
                        st.rerun()
                col_idx += 1
        if next_unlockable and col_idx < 5:
            with btn_cols[col_idx]:
                if st.button(f"🔓 Unlock {next_unlockable}", key=f"unlock_{next_unlockable}", use_container_width=True):
                    st.session_state.unlocked_days.append(next_unlockable)
                    save_session_to_shared()
                    st.rerun()
        
        day_tabs = st.tabs([f"{day} ({week_dates.get(day, '')})" for day in visible_days])
        for day_tab, day in zip(day_tabs, visible_days):
            with day_tab:
                day_idx_loop = DAY_ABBR_ORDER.index(day) if day in DAY_ABBR_ORDER else -1
                if day_idx_loop < today_idx:
                    st.warning(f"Carryover from {day}")
                if day_idx_loop > today_idx:
                    st.info(f"{day} — unlocked for early prep")
                day_state = st.session_state.cycle_team_tracking.get(day, {})
                day_facs = cycle_facilities.get(day, [])
                done = sum(1 for f in day_facs if stage_counts(day_state.get(f, {}))[0] == len(CYCLE_STAGE_ORDER))
                st.markdown(f"**{day}**: {done} of {len(day_facs)} complete")
                for facility in day_facs:
                    if facility not in day_state:
                        day_state[facility] = {s: "" for s in CYCLE_STAGE_ORDER}
                    render_cycle_facility(day, facility, day_state[facility])
    
    # ============ TAB 3: HIGH DOLLAR TRACKING ============
    with cycle_tab3:
        dollar_facilities = _cycle_facilities
        
        for day, facs in dollar_facilities.items():
            if day not in st.session_state.dollar_tracking:
                st.session_state.dollar_tracking[day] = {}
            for fac in facs:
                if fac not in st.session_state.dollar_tracking[day]:
                    st.session_state.dollar_tracking[day][fac] = {stage: "" for stage in CYCLE_DOLLAR_STAGE_ORDER}
        
        st.markdown("### Cycle High Dollar Tracking")
        today_abbr = _today_abbr()
        week_dates = get_week_dates()
        today_idx = DAY_ABBR_ORDER.index(today_abbr) if today_abbr in DAY_ABBR_ORDER else -1
        week_num = get_current_week_number()
        
        st.caption(f"High-dollar billing summary as of {datetime.now().strftime('%A %Y-%m-%d %H:%M')} (Week {week_num})")
        
        # Dashboard summary - matching Cycle Team Dashboard style
        if today_abbr in dollar_facilities and dollar_facilities[today_abbr]:
            today_facs = dollar_facilities[today_abbr]
            dollar_rows = []
            for fac in today_facs:
                dm = st.session_state.dollar_tracking.get(today_abbr, {}).get(fac, {})
                if dm.get("_no_meds"):
                    status = "No Meds"
                    pct = 100
                else:
                    c, t = stage_counts(dm, CYCLE_DOLLAR_STAGE_ORDER)
                    pct = int(c / t * 100) if t else 0
                    if c == 0:
                        status = "Not Started"
                    elif c >= t:
                        status = "Completed"
                    else:
                        status = "In Progress"
                dollar_rows.append({
                    "Facility": fac,
                    "Progress": f"{c}/{t} ({pct}%)" if not dm.get("_no_meds") else "N/A",
                    "Status": status
                })
            
            # Summary metrics matching Cycle Team Dashboard
            total = len(dollar_rows)
            completed = sum(1 for r in dollar_rows if r["Status"] in ["Completed", "No Meds"])
            in_progress = sum(1 for r in dollar_rows if r["Status"] == "In Progress")
            not_started = sum(1 for r in dollar_rows if r["Status"] == "Not Started")
            
            m1, m2, m3, m4 = st.columns(4)
            m1.metric("Total Facilities", total)
            m2.metric("Completed", completed)
            m3.metric("In Progress", in_progress)
            m4.metric("Not Started", not_started)
            
            st.dataframe(pd.DataFrame(dollar_rows), use_container_width=True, hide_index=True)
        
        st.divider()
        st.markdown("#### Tracking Details")
        st.caption("Click a facility $$ to expand and mark stages complete.")
        
        def get_visible_days_dollar(ts, fc):
            if today_abbr not in DAY_ABBR_ORDER:
                return DAY_ABBR_ORDER
            visible = []
            for i, d in enumerate(DAY_ABBR_ORDER):
                if i == today_idx:
                    visible.append(d)
                elif i < today_idx:
                    for f in fc.get(d, []):
                        fs = ts.get(d, {}).get(f, {})
                        if fs.get("_no_meds"):
                            continue
                        c, t = stage_counts(fs, CYCLE_DOLLAR_STAGE_ORDER)
                        if c < t:
                            visible.append(d)
                            break
            return visible
        
        visible_days = get_visible_days_dollar(st.session_state.dollar_tracking, dollar_facilities)
        if not visible_days:
            visible_days = [today_abbr] if today_abbr in DAY_ABBR_ORDER else DAY_ABBR_ORDER
        
        for unlocked in st.session_state.dollar_unlocked_days:
            if unlocked not in visible_days and unlocked in DAY_ABBR_ORDER:
                visible_days.append(unlocked)
        visible_days = sorted(set(visible_days), key=lambda d: DAY_ABBR_ORDER.index(d) if d in DAY_ABBR_ORDER else 99)
        
        future_days = DAY_ABBR_ORDER[today_idx + 1:] if today_idx >= 0 else []
        next_unlockable = next((fd for fd in future_days if fd not in st.session_state.dollar_unlocked_days), None)
        
        btn_cols = st.columns(5)
        col_idx = 0
        for unlocked in st.session_state.dollar_unlocked_days:
            if col_idx < 5:
                with btn_cols[col_idx]:
                    if st.button(f"🔒 Lock {unlocked}", key=f"$relock_{unlocked}", use_container_width=True):
                        st.session_state.dollar_unlocked_days.remove(unlocked)
                        st.session_state.dollar_unlocked_days = [d for d in st.session_state.dollar_unlocked_days if DAY_ABBR_ORDER.index(d) < DAY_ABBR_ORDER.index(unlocked)]
                        save_session_to_shared()
                        st.rerun()
                col_idx += 1
        if next_unlockable and col_idx < 5:
            with btn_cols[col_idx]:
                if st.button(f"🔓 Unlock {next_unlockable}", key=f"$unlock_{next_unlockable}", use_container_width=True):
                    st.session_state.dollar_unlocked_days.append(next_unlockable)
                    save_session_to_shared()
                    st.rerun()
        
        day_tabs = st.tabs([f"{day} $$ ({week_dates.get(day, '')})" for day in visible_days])
        for day_tab, day in zip(day_tabs, visible_days):
            with day_tab:
                day_idx_loop = DAY_ABBR_ORDER.index(day) if day in DAY_ABBR_ORDER else -1
                if day_idx_loop < today_idx:
                    st.warning(f"Carryover from {day}")
                if day_idx_loop > today_idx:
                    st.info(f"{day} $$ — unlocked for early prep")
                day_state = st.session_state.dollar_tracking.get(day, {})
                day_facs = dollar_facilities.get(day, [])
                done = sum(1 for f in day_facs if day_state.get(f, {}).get("_no_meds") or stage_counts(day_state.get(f, {}), CYCLE_DOLLAR_STAGE_ORDER)[0] == len(CYCLE_DOLLAR_STAGE_ORDER))
                st.markdown(f"**{day} $$**: {done} of {len(day_facs)} complete")
                for facility in day_facs:
                    if facility not in day_state:
                        day_state[facility] = {s: "" for s in CYCLE_DOLLAR_STAGE_ORDER}
                    render_dollar_facility(day, facility, day_state[facility])

    # ============ TAB 4: CYCLE BAG COUNT ============
    with cycle_tab4:
        import uuid
        
        # Load bag count state
        bag_state = supa.load_bag_count_state()
        if "bag_batches" not in st.session_state:
            st.session_state.bag_batches = bag_state.get("batches", {})
        if "bag_counts" not in st.session_state:
            st.session_state.bag_counts = bag_state.get("counts", {})
        if "bag_unlocked_days" not in st.session_state:
            st.session_state.bag_unlocked_days = bag_state.get("unlocked_days", [])
        if "bag_completed_days" not in st.session_state:
            st.session_state.bag_completed_days = bag_state.get("completed_days", [])
        
        st.markdown("### Cycle Bag Count")
        st.caption("Track bag counts and census per batch for each facility.")
        
        today_abbr = _today_abbr()
        week_dates = get_week_dates()
        today_idx = DAY_ABBR_ORDER.index(today_abbr) if today_abbr in DAY_ABBR_ORDER else -1
        
        # Determine visible days (entire week Mon-Fri unless marked complete + unlocked future)
        def get_visible_days_bag():
            if today_abbr not in DAY_ABBR_ORDER:
                return DAY_ABBR_ORDER
            visible = []
            completed = st.session_state.bag_completed_days
            for i, d in enumerate(DAY_ABBR_ORDER):
                # Show all days Mon-Fri unless marked complete
                if d not in completed:
                    visible.append(d)
            return visible
        
        visible_days = get_visible_days_bag()
        for unlocked in st.session_state.bag_unlocked_days:
            if unlocked not in visible_days and unlocked in DAY_ABBR_ORDER:
                visible_days.append(unlocked)
        visible_days = sorted(set(visible_days), key=lambda d: DAY_ABBR_ORDER.index(d) if d in DAY_ABBR_ORDER else 99)
        
        future_days = DAY_ABBR_ORDER[today_idx + 1:] if today_idx >= 0 else []
        next_unlockable = next((fd for fd in future_days if fd not in st.session_state.bag_unlocked_days), None)
        
        # Day Complete / Lock/Unlock buttons
        st.markdown("#### Day Controls")
        btn_cols = st.columns(5)
        col_idx = 0
        
        # Show "Day Complete" buttons for today and past days that are visible
        completable_days = [d for d in visible_days if DAY_ABBR_ORDER.index(d) <= today_idx]
        for day in completable_days:
            if col_idx < 5:
                with btn_cols[col_idx]:
                    if st.button(f"✅ {day} ({week_dates.get(day, '')}) Complete", key=f"bag_complete_{day}", use_container_width=True):
                        if day not in st.session_state.bag_completed_days:
                            st.session_state.bag_completed_days.append(day)
                        supa.save_bag_count_state({
                            "batches": st.session_state.bag_batches,
                            "counts": st.session_state.bag_counts,
                            "unlocked_days": st.session_state.bag_unlocked_days,
                            "completed_days": st.session_state.bag_completed_days,
                        })
                        st.rerun()
                col_idx += 1
        
        # Show unlock button for future days
        if next_unlockable and col_idx < 5:
            with btn_cols[col_idx]:
                if st.button(f"🔓 Unlock {next_unlockable} ({week_dates.get(next_unlockable, '')})", key=f"bag_unlock_{next_unlockable}", use_container_width=True):
                    st.session_state.bag_unlocked_days.append(next_unlockable)
                    supa.save_bag_count_state({
                        "batches": st.session_state.bag_batches,
                        "counts": st.session_state.bag_counts,
                        "unlocked_days": st.session_state.bag_unlocked_days,
                        "completed_days": st.session_state.bag_completed_days,
                    })
                    st.rerun()
                col_idx += 1
        
        # Show lock buttons for unlocked future days
        for unlocked in st.session_state.bag_unlocked_days:
            if col_idx < 5:
                with btn_cols[col_idx]:
                    if st.button(f"🔒 Lock {unlocked}", key=f"bag_relock_{unlocked}", use_container_width=True):
                        st.session_state.bag_unlocked_days.remove(unlocked)
                        st.session_state.bag_unlocked_days = [d for d in st.session_state.bag_unlocked_days if DAY_ABBR_ORDER.index(d) < DAY_ABBR_ORDER.index(unlocked)]
                        supa.save_bag_count_state({
                            "batches": st.session_state.bag_batches,
                            "counts": st.session_state.bag_counts,
                            "unlocked_days": st.session_state.bag_unlocked_days,
                            "completed_days": st.session_state.bag_completed_days,
                        })
                        st.rerun()
                col_idx += 1
        
        # Manual save button (in case auto-save failed)
        save_col1, save_col2 = st.columns([1, 3])
        with save_col1:
            if st.button("💾 Save All to Database", key="manual_save_bags", use_container_width=True):
                supa.save_bag_count_state({
                    "batches": st.session_state.bag_batches,
                    "counts": st.session_state.bag_counts,
                    "unlocked_days": st.session_state.bag_unlocked_days,
                    "completed_days": st.session_state.bag_completed_days,
                })
                st.success("✅ Saved to database!")
        with save_col2:
            if supa.using_supabase():
                st.caption("🟢 Connected to Supabase")
            else:
                st.caption("🔴 Supabase not connected — data in session only")
        
        # Batch Management in expander
        with st.expander("⚙️ Manage Batches (per facility)", expanded=False):
            st.caption("Each facility has its own batches. Add/remove batches here.")
            
            bag_facilities = _cycle_facilities
            all_facs = []
            for day_facs in bag_facilities.values():
                all_facs.extend(day_facs)
            all_facs = sorted(set(all_facs))
            
            if all_facs:
                selected_fac = st.selectbox("Select Facility", all_facs, key="batch_mgmt_fac")
                
                # Show current batches for this facility
                current_batches = st.session_state.bag_batches.get(selected_fac, [])
                
                if current_batches:
                    st.markdown(f"**Current batches for {selected_fac}:**")
                    for i, batch in enumerate(current_batches):
                        bc1, bc2 = st.columns([3, 1])
                        with bc1:
                            st.text(f"• {batch['name']}")
                        with bc2:
                            if st.button("🗑️", key=f"del_batch_{selected_fac}_{batch['id']}", help="Remove batch"):
                                st.session_state.bag_batches[selected_fac] = [b for b in current_batches if b['id'] != batch['id']]
                                supa.save_bag_count_state({
                                    "batches": st.session_state.bag_batches,
                                    "counts": st.session_state.bag_counts,
                                    "unlocked_days": st.session_state.bag_unlocked_days,
                                    "completed_days": st.session_state.bag_completed_days,
                                })
                                st.rerun()
                else:
                    st.info(f"No batches defined for {selected_fac} yet.")
                
                # Add new batch
                with st.form(f"add_batch_form_{selected_fac}", clear_on_submit=True):
                    new_batch_name = st.text_input("New Batch Name", placeholder="e.g., A Wing, Main Building")
                    if st.form_submit_button("➕ Add Batch", use_container_width=True):
                        if new_batch_name.strip():
                            if selected_fac not in st.session_state.bag_batches:
                                st.session_state.bag_batches[selected_fac] = []
                            st.session_state.bag_batches[selected_fac].append({
                                "name": new_batch_name.strip(),
                                "id": str(uuid.uuid4())[:8]
                            })
                            supa.save_bag_count_state({
                                "batches": st.session_state.bag_batches,
                                "counts": st.session_state.bag_counts,
                                "unlocked_days": st.session_state.bag_unlocked_days,
                                "completed_days": st.session_state.bag_completed_days,
                            })
                            st.rerun()
        
        st.divider()
        
        # Day tabs for bag count entry
        if visible_days:
            day_tabs = st.tabs([f"📦 {day} ({week_dates.get(day, '')})" for day in visible_days])
            for day_tab, day in zip(day_tabs, visible_days):
                with day_tab:
                    day_idx_loop = DAY_ABBR_ORDER.index(day) if day in DAY_ABBR_ORDER else -1
                    if day_idx_loop < today_idx - 2:
                        st.warning(f"Viewing older day: {day}")
                    if day_idx_loop > today_idx:
                        st.info(f"{day} — unlocked for early prep")
                    
                    day_facs = _cycle_facilities.get(day, [])
                    
                    if not day_facs:
                        st.info(f"No facilities scheduled for {day}")
                        continue
                    
                    # Initialize day in counts if needed
                    if day not in st.session_state.bag_counts:
                        st.session_state.bag_counts[day] = {}
                    
                    # Calculate totals for the day
                    day_total_bags = 0
                    day_total_census = 0
                    
                    for facility in day_facs:
                        batches = st.session_state.bag_batches.get(facility, [])
                        
                        if facility not in st.session_state.bag_counts[day]:
                            st.session_state.bag_counts[day][facility] = {}
                        
                        fac_counts = st.session_state.bag_counts[day][facility]
                        
                        # Calculate facility totals
                        fac_total_bags = sum(fac_counts.get(b['id'], {}).get('bags', 0) or 0 for b in batches)
                        fac_total_census = sum(fac_counts.get(b['id'], {}).get('census', 0) or 0 for b in batches)
                        day_total_bags += fac_total_bags
                        day_total_census += fac_total_census
                        
                        with st.expander(f"**{facility}** — {len(batches)} batch(es) | Bags: {fac_total_bags} | Census: {fac_total_census}", expanded=False):
                            if not batches:
                                st.warning(f"No batches defined for {facility}. Add batches in 'Manage Batches' above.")
                            else:
                                # Column headers
                                hdr1, hdr2, hdr3 = st.columns([2, 1, 1])
                                with hdr1:
                                    st.caption("Batch")
                                with hdr2:
                                    st.caption("Bags")
                                with hdr3:
                                    st.caption("Census")
                                
                                # Create a grid for batch entry
                                for batch in batches:
                                    batch_id = batch['id']
                                    batch_name = batch['name']
                                    
                                    if batch_id not in fac_counts:
                                        fac_counts[batch_id] = {'bags': None, 'census': None}
                                    
                                    col1, col2, col3 = st.columns([2, 1, 1])
                                    with col1:
                                        st.markdown(f"**{batch_name}**")
                                    with col2:
                                        new_bags = st.number_input(
                                            "Bags",
                                            min_value=0,
                                            value=fac_counts[batch_id].get('bags') or 0,
                                            key=f"bags_{day}_{facility}_{batch_id}",
                                            label_visibility="collapsed"
                                        )
                                        if new_bags != fac_counts[batch_id].get('bags'):
                                            fac_counts[batch_id]['bags'] = new_bags
                                            supa.save_bag_count_state({
                                                "batches": st.session_state.bag_batches,
                                                "counts": st.session_state.bag_counts,
                                                "unlocked_days": st.session_state.bag_unlocked_days,
                                                "completed_days": st.session_state.bag_completed_days,
                                            })
                                    with col3:
                                        new_census = st.number_input(
                                            "Census",
                                            min_value=0,
                                            value=fac_counts[batch_id].get('census') or 0,
                                            key=f"census_{day}_{facility}_{batch_id}",
                                            label_visibility="collapsed"
                                        )
                                        if new_census != fac_counts[batch_id].get('census'):
                                            fac_counts[batch_id]['census'] = new_census
                                            supa.save_bag_count_state({
                                                "batches": st.session_state.bag_batches,
                                                "counts": st.session_state.bag_counts,
                                                "unlocked_days": st.session_state.bag_unlocked_days,
                                                "completed_days": st.session_state.bag_completed_days,
                                            })
                                
                                # Show facility totals
                                st.markdown(f"**Total: {fac_total_bags} bags, {fac_total_census} census**")
                    
                    # Day summary
                    st.divider()
                    m1, m2, m3 = st.columns(3)
                    m1.metric("Facilities", len(day_facs))
                    m2.metric("Total Bags", day_total_bags)
                    m3.metric("Total Census", day_total_census)

# --- PAGE: Facility Management ---
if current_page == "Facility Management":
    st.markdown("### Facility Management")
    
    # Load current config
    current_config = load_facilities_config()
    week_num = get_current_week_number()
    today_str = datetime.now().strftime("%Y-%m-%d")
    
    # Two columns: Add facility and Admin Override
    add_col, override_col = st.columns(2)
    
    with add_col:
        st.markdown("#### ➕ Add Facility")
        with st.form("add_facility_form", clear_on_submit=True):
            new_facility_name = st.text_input("Name", placeholder="e.g., Sunny Acres")
            c1, c2, c3 = st.columns(3)
            with c1:
                new_facility_day = st.selectbox("Day", DAY_ABBR_ORDER)
            with c2:
                new_facility_freq = st.selectbox("Freq", FREQUENCY_OPTIONS)
            with c3:
                new_start_date = st.date_input("Start", value=datetime.now())
            
            if st.form_submit_button("Add", use_container_width=True):
                if new_facility_name.strip():
                    fname = new_facility_name.strip()
                    existing_names = [f["name"] for f in current_config.get(new_facility_day, [])]
                    if fname not in existing_names:
                        if new_facility_day not in current_config:
                            current_config[new_facility_day] = []
                        new_fac = {"name": fname, "frequency": new_facility_freq}
                        if new_facility_freq != "Weekly":
                            new_fac["start_date"] = new_start_date.strftime("%Y-%m-%d")
                        current_config[new_facility_day].append(new_fac)
                        save_facilities_config(current_config)
                        st.rerun()
    
    with override_col:
        st.markdown("#### 🔐 Admin Override")
        st.caption("Mark all facilities complete or undo")
        
        # Track last override for undo
        if "last_override" not in st.session_state:
            st.session_state.last_override = None
        
        with st.form("admin_override_form"):
            override_date = st.date_input("Date", value=datetime.now() - timedelta(days=1))
            override_day = override_date.strftime("%a")[:3]
            override_type = st.radio("Apply to:", ["Cycle Tracking", f"{DOLLAR_DISPLAY} Tracking", "Both"], horizontal=True)
            
            c1, c2 = st.columns(2)
            with c1:
                do_complete = st.form_submit_button("✓ Mark Complete", use_container_width=True)
            with c2:
                do_undo = st.form_submit_button("↩ Undo/Clear", use_container_width=True)
            
            if do_complete or do_undo:
                override_date_str = override_date.strftime("%Y-%m-%d")
                facs_for_day = [f["name"] for f in current_config.get(override_day, [])]
                
                if do_complete:
                    # Mark all complete
                    if override_type in ["Cycle Tracking", "Both"]:
                        if "cycle_team_tracking" not in st.session_state:
                            st.session_state.cycle_team_tracking = {}
                        if override_day not in st.session_state.cycle_team_tracking:
                            st.session_state.cycle_team_tracking[override_day] = {}
                        for fac in facs_for_day:
                            st.session_state.cycle_team_tracking[override_day][fac] = {
                                stage: "ADMIN" for stage in CYCLE_STAGE_ORDER
                            }
                            log_stage_to_excel(fac, "ADMIN OVERRIDE", f"ALL ({override_date_str})")
                    
                    if override_type in [f"{DOLLAR_DISPLAY} Tracking", "Both"]:
                        if "dollar_tracking" not in st.session_state:
                            st.session_state.dollar_tracking = {}
                        if override_day not in st.session_state.dollar_tracking:
                            st.session_state.dollar_tracking[override_day] = {}
                        for fac in facs_for_day:
                            st.session_state.dollar_tracking[override_day][fac] = {
                                stage: "ADMIN" for stage in CYCLE_DOLLAR_STAGE_ORDER
                            }
                            log_stage_to_excel(f"{fac} $$", "ADMIN OVERRIDE", f"ALL ({override_date_str})")
                    
                    st.session_state.last_override = {"day": override_day, "type": override_type, "date": override_date_str}
                    st.success(f"Marked {len(facs_for_day)} facilities complete for {override_day}")
                
                elif do_undo:
                    # Clear/reset the day
                    if override_type in ["Cycle Tracking", "Both"]:
                        if override_day in st.session_state.get("cycle_team_tracking", {}):
                            for fac in facs_for_day:
                                st.session_state.cycle_team_tracking[override_day][fac] = {
                                    stage: "" for stage in CYCLE_STAGE_ORDER
                                }
                            log_stage_to_excel(f"{override_day} ALL", "UNDO/CLEAR", override_date_str)
                    
                    if override_type in [f"{DOLLAR_DISPLAY} Tracking", "Both"]:
                        if override_day in st.session_state.get("dollar_tracking", {}):
                            for fac in facs_for_day:
                                st.session_state.dollar_tracking[override_day][fac] = {
                                    stage: "" for stage in CYCLE_DOLLAR_STAGE_ORDER
                                }
                            log_stage_to_excel(f"{override_day} ALL $$", "UNDO/CLEAR", override_date_str)
                    
                    st.session_state.last_override = None
                    st.warning(f"Cleared {len(facs_for_day)} facilities for {override_day}")
                
                # Persist to Supabase/shared state
                save_shared_state({
                    "cycle_team_tracking": st.session_state.get("cycle_team_tracking", {}),
                    "dollar_tracking": st.session_state.get("dollar_tracking", {}),
                    "unlocked_days": st.session_state.get("unlocked_days", []),
                    "dollar_unlocked_days": st.session_state.get("dollar_unlocked_days", []),
                })
                st.rerun()
    
    st.divider()
    
    # Compact frequency legend
    st.caption(f"Week {week_num} · ✓=active · Weekly | Every 2wk (from start) | Every 4wk (from start)")
    
    # Edit/Remove facilities - compact view
    for day in DAY_ABBR_ORDER:
        day_facilities = current_config.get(day, [])
        with st.expander(f"**{day}** ({len(day_facilities)})", expanded=False):
            if not day_facilities:
                st.caption("No facilities.")
            else:
                for i, fac_data in enumerate(list(day_facilities)):
                    fac_name = fac_data["name"]
                    fac_freq = fac_data.get("frequency", "Weekly")
                    fac_start = fac_data.get("start_date")
                    is_active = facility_active_this_week(fac_freq, fac_start)
                    ind = "✓" if is_active else "○"
                    
                    # Single compact row
                    cols = st.columns([0.3, 2, 1, 1, 0.8, 0.4])
                    with cols[0]:
                        st.write(ind)
                    with cols[1]:
                        st.write(fac_name)
                    with cols[2]:
                        new_freq = st.selectbox("F", FREQUENCY_OPTIONS, 
                            index=FREQUENCY_OPTIONS.index(fac_freq) if fac_freq in FREQUENCY_OPTIONS else 0,
                            key=f"f_{day}_{i}", label_visibility="collapsed")
                        if new_freq != fac_freq:
                            current_config[day][i]["frequency"] = new_freq
                            if new_freq != "Weekly" and not fac_start:
                                current_config[day][i]["start_date"] = today_str
                            save_facilities_config(current_config)
                            st.rerun()
                    with cols[3]:
                        if fac_freq != "Weekly":
                            try:
                                cur_date = datetime.strptime(fac_start, "%Y-%m-%d").date() if fac_start else datetime.now().date()
                            except:
                                cur_date = datetime.now().date()
                            new_date = st.date_input("D", value=cur_date, key=f"d_{day}_{i}", label_visibility="collapsed")
                            if new_date.strftime("%Y-%m-%d") != fac_start:
                                current_config[day][i]["start_date"] = new_date.strftime("%Y-%m-%d")
                                save_facilities_config(current_config)
                                st.rerun()
                        else:
                            st.caption("—")
                    with cols[4]:
                        move_opts = ["—"] + [d for d in DAY_ABBR_ORDER if d != day]
                        move_to = st.selectbox("M", move_opts, key=f"m_{day}_{i}", label_visibility="collapsed")
                        if move_to != "—":
                            fac_to_move = current_config[day].pop(i)
                            if move_to not in current_config:
                                current_config[move_to] = []
                            current_config[move_to].append(fac_to_move)
                            save_facilities_config(current_config)
                            st.rerun()
                    with cols[5]:
                        if st.button("🗑", key=f"x_{day}_{i}"):
                            current_config[day].pop(i)
                            save_facilities_config(current_config)
                            st.rerun()

# --- PAGE: Pharmacy Management ---
if current_page == "Pharmacy Management":
    st.markdown("### 🏥 Pharmacy Management")
    
    # Load master facilities
    if "master_facilities" not in st.session_state:
        st.session_state.master_facilities = supa.load_master_facilities()
    
    master_facs = st.session_state.master_facilities
    
    def calculate_next_renewal(start_date_str: str, original_term: int, renewal_term: int) -> tuple[str, str]:
        """Calculate next renewal date and whether it's original or renewal term."""
        try:
            start = datetime.strptime(start_date_str, "%Y-%m-%d")
        except (ValueError, TypeError):
            return ("Invalid date", "Unknown")
        
        today = datetime.now()
        original_end = start + relativedelta(years=original_term)
        if today < original_end:
            return (original_end.strftime("%b %d, %Y"), f"Original ({original_term}yr)")
        
        current_end = original_end
        while current_end <= today:
            current_end += relativedelta(years=renewal_term)
        
        return (current_end.strftime("%b %d, %Y"), f"Renewal ({renewal_term}yr)")
    
    # Tabs for different sections
    pharm_tab1, = st.tabs(["📋 Facility Directory"])
    
    with pharm_tab1:
        # Add New Facility button that expands to form
        with st.expander("➕ Add New Facility", expanded=False):
            with st.form("add_facility_form", clear_on_submit=True):
                col1, col2 = st.columns(2)
                with col1:
                    new_name = st.text_input("Facility Name", placeholder="e.g., Sunrise Manor")
                    new_start = st.date_input("Contract Start Date")
                with col2:
                    new_original = st.number_input("Original Term (years)", min_value=1, max_value=10, value=3)
                    new_renewal = st.number_input("Renewal Term (years)", min_value=1, max_value=10, value=1)
                
                if st.form_submit_button("➕ Add Facility", use_container_width=True):
                    if new_name.strip():
                        existing_names = [f["name"].lower() for f in master_facs]
                        if new_name.strip().lower() in existing_names:
                            st.error(f"Facility '{new_name}' already exists")
                        else:
                            master_facs.append({
                                "name": new_name.strip(),
                                "start_date": new_start.strftime("%Y-%m-%d"),
                                "original_term": new_original,
                                "renewal_term": new_renewal,
                            })
                            supa.save_master_facilities(master_facs)
                            st.success(f"Added '{new_name}'")
                            st.rerun()
                    else:
                        st.warning("Please enter a facility name")
        
        # Display facility list
        st.markdown(f"**{len(master_facs)} facilities**")
        
        if master_facs:
            # Build display data (sorted alphabetically)
            display_rows = []
            for fac in sorted(master_facs, key=lambda x: x["name"].lower()):
                next_renewal, term_type = calculate_next_renewal(
                    fac.get("start_date", ""),
                    fac.get("original_term", 3),
                    fac.get("renewal_term", 1)
                )
                display_rows.append({
                    "Facility": fac["name"],
                    "Start Date": fac.get("start_date", "N/A"),
                    "Original Term": f"{fac.get('original_term', 'N/A')} yr",
                    "Renewal Term": f"{fac.get('renewal_term', 'N/A')} yr",
                    "Next Renewal": next_renewal,
                    "Term Type": term_type,
                })
            
            st.dataframe(pd.DataFrame(display_rows), use_container_width=True, hide_index=True)
            
            # Edit/Delete - collapsed by default (sorted alphabetically)
            with st.expander("✏️ Edit or Remove Facilities", expanded=False):
                sorted_facs = sorted(master_facs, key=lambda x: x["name"].lower())
                for fac in sorted_facs:
                    # Find actual index in master_facs for editing
                    actual_idx = next(j for j, f in enumerate(master_facs) if f["name"] == fac["name"])
                    with st.expander(f"**{fac['name']}**"):
                        edit_col1, edit_col2, edit_col3 = st.columns([2, 2, 1])
                        with edit_col1:
                            edit_start = st.date_input(
                                "Start Date", 
                                value=datetime.strptime(fac.get("start_date", "2020-01-01"), "%Y-%m-%d"),
                                key=f"edit_start_{fac['name']}"
                            )
                            edit_original = st.number_input(
                                "Original Term (yr)", 
                                min_value=1, max_value=10, 
                                value=fac.get("original_term", 3),
                                key=f"edit_orig_{fac['name']}"
                            )
                        with edit_col2:
                            edit_renewal = st.number_input(
                                "Renewal Term (yr)", 
                                min_value=1, max_value=10, 
                                value=fac.get("renewal_term", 1),
                                key=f"edit_renew_{fac['name']}"
                            )
                            if st.button("💾 Save Changes", key=f"save_{fac['name']}", use_container_width=True):
                                master_facs[actual_idx]["start_date"] = edit_start.strftime("%Y-%m-%d")
                                master_facs[actual_idx]["original_term"] = edit_original
                                master_facs[actual_idx]["renewal_term"] = edit_renewal
                                supa.save_master_facilities(master_facs)
                                st.success("Saved!")
                                st.rerun()
                        with edit_col3:
                            st.write("")
                            st.write("")
                            if st.button("🗑️ Delete", key=f"del_{fac['name']}", type="secondary"):
                                del master_facs[actual_idx]
                                supa.save_master_facilities(master_facs)
                                st.rerun()
        else:
            st.info("No facilities added yet. Click 'Add New Facility' above to add your first facility.")

# --- PAGE: QA ---
if current_page == "QA":
    st.title("QA")
    
    # Load master facilities for dropdown
    if "master_facilities" not in st.session_state:
        st.session_state.master_facilities = supa.load_master_facilities()
    master_facs = st.session_state.master_facilities
    facility_names_list = sorted([f["name"] for f in master_facs])
    
    # Load BNDD licenses
    if "bndd_licenses" not in st.session_state:
        st.session_state.bndd_licenses = supa.load_bndd_licenses()
    bndd_licenses = st.session_state.bndd_licenses
    
    # Load Cubex re-stock data
    if "cubex_restock" not in st.session_state:
        st.session_state.cubex_restock = supa.load_cubex_restock()
    cubex_restock = st.session_state.cubex_restock
    
    # Tabs: Dashboard, BNDD License, Cubex Re-Stock
    qa_tab1, qa_tab2, qa_tab3 = st.tabs(["📊 Dashboard", "📋 BNDD License", "📦 Cubex Re-Stock"])
    
    # ============ TAB 1: QA Dashboard ============
    with qa_tab1:
        st.markdown("### QA Dashboard")
        
        # BNDD Licenses expiring within 90 days
        st.markdown("#### 📋 BNDD Licenses Expiring Soon")
        
        today = datetime.now()
        expiring_licenses = []
        
        for lic in bndd_licenses:
            try:
                exp_date = datetime.strptime(lic["expiration_date"], "%Y-%m-%d")
                days_until = (exp_date - today).days
                if days_until <= 90:
                    expiring_licenses.append({
                        "Facility": lic["facility"],
                        "License #": lic.get("license_number", "N/A"),
                        "Expiration Date": exp_date.strftime("%b %d, %Y"),
                        "Days Until Expiration": days_until,
                    })
            except:
                pass
        
        if expiring_licenses:
            # Sort by soonest expiration first
            expiring_licenses.sort(key=lambda x: x["Days Until Expiration"])
            
            # Create DataFrame
            df = pd.DataFrame(expiring_licenses)
            
            # Style function for row colors
            def color_rows(row):
                days = row["Days Until Expiration"]
                if days <= 30:
                    return ["background-color: #fecaca"] * len(row)  # Red
                elif days <= 60:
                    return ["background-color: #fef08a"] * len(row)  # Yellow
                else:
                    return ["background-color: #bbf7d0"] * len(row)  # Green
            
            # Apply styling
            styled_df = df.style.apply(color_rows, axis=1)
            
            # Display
            st.dataframe(styled_df, use_container_width=True, hide_index=True)
            st.caption(f"{len(expiring_licenses)} license(s) expiring within 90 days")
        else:
            st.success("✅ No BNDD licenses expiring within 90 days")
        
        st.divider()
        
        # Cubex Re-Stocks due within 90 days
        st.markdown("#### 📦 Cubex Re-Stocks Due Soon")
        
        expiring_cubex = []
        for entry in cubex_restock:
            try:
                next_due_dt = datetime.strptime(entry["next_restock_due"], "%Y-%m-%d")
                days_until = (next_due_dt - today).days
                if days_until <= 90:
                    expiring_cubex.append({
                        "Facility": entry["facility"],
                        "Serial Number": entry.get("serial_number", "") or "—",
                        "Next Re-Stock Due": next_due_dt.strftime("%b %d, %Y"),
                        "Days Until Due": days_until,
                    })
            except:
                pass
        
        if expiring_cubex:
            # Sort by soonest due first
            expiring_cubex.sort(key=lambda x: x["Days Until Due"])
            
            # Create DataFrame
            cubex_df = pd.DataFrame(expiring_cubex)
            
            # Style function for row colors
            def color_cubex_dash_rows(row):
                days = row["Days Until Due"]
                if days <= 30:
                    return ["background-color: #fecaca"] * len(row)  # Red
                elif days <= 60:
                    return ["background-color: #fef08a"] * len(row)  # Yellow
                else:
                    return ["background-color: #bbf7d0"] * len(row)  # Green
            
            # Apply styling
            styled_cubex_df = cubex_df.style.apply(color_cubex_dash_rows, axis=1)
            
            # Display
            st.dataframe(styled_cubex_df, use_container_width=True, hide_index=True)
            st.caption(f"{len(expiring_cubex)} Cubex machine(s) due for re-stock within 90 days")
        else:
            st.success("✅ No Cubex re-stocks due within 90 days")
    
    # ============ TAB 2: BNDD License ============
    with qa_tab2:
        st.markdown("### BNDD License Tracking")
        st.caption("Track facility BNDD licenses and expiration dates.")
        
        # Add New License
        with st.expander("➕ Add New License", expanded=False):
            with st.form("add_bndd_form", clear_on_submit=True):
                if facility_names_list:
                    # Filter out facilities that already have a license
                    existing_facilities = {lic["facility"] for lic in bndd_licenses}
                    available_facilities = [f for f in facility_names_list if f not in existing_facilities]
                    
                    if available_facilities:
                        col1, col2, col3 = st.columns([2, 2, 2])
                        with col1:
                            new_facility = st.selectbox("Facility", available_facilities)
                        with col2:
                            new_license_num = st.text_input("License Number", placeholder="e.g., 12345-BNDD")
                        with col3:
                            new_exp_date = st.date_input("Expiration Date")
                        
                        if st.form_submit_button("➕ Add License", use_container_width=True):
                            if new_facility and new_license_num.strip():
                                bndd_licenses.append({
                                    "facility": new_facility,
                                    "license_number": new_license_num.strip(),
                                    "expiration_date": new_exp_date.strftime("%Y-%m-%d"),
                                })
                                supa.save_bndd_licenses(bndd_licenses)
                                st.session_state.bndd_licenses = bndd_licenses
                                st.success(f"Added license for {new_facility}")
                                st.rerun()
                            else:
                                st.warning("Please fill in all fields")
                    else:
                        st.info("All facilities already have licenses assigned.")
                        st.form_submit_button("➕ Add License", disabled=True)
                else:
                    st.warning("No facilities found. Add facilities in Pharmacy Management first.")
                    st.form_submit_button("➕ Add License", disabled=True)
        
        # Display licenses table
        if bndd_licenses:
            st.markdown(f"**{len(bndd_licenses)} licenses tracked**")
            
            # Sort alphabetically by facility name
            sorted_licenses = sorted(bndd_licenses, key=lambda x: x.get("facility", "").lower())
            
            today = datetime.now()
            
            # Table header with custom styling
            st.markdown("""
            <style>
            .bndd-header {
                display: flex;
                font-weight: 700;
                padding: 12px 16px;
                background: #f1f5f9;
                border-radius: 8px 8px 0 0;
                border: 1px solid #e2e8f0;
                margin-top: 16px;
            }
            .bndd-header > div:nth-child(1) { flex: 2; }
            .bndd-header > div:nth-child(2) { flex: 2; }
            .bndd-header > div:nth-child(3) { flex: 1.5; }
            .bndd-header > div:nth-child(4) { flex: 1; }
            .bndd-row {
                display: flex;
                padding: 12px 16px;
                border-left: 1px solid #e2e8f0;
                border-right: 1px solid #e2e8f0;
                border-bottom: 1px solid #e2e8f0;
                align-items: center;
            }
            .bndd-row:last-child { border-radius: 0 0 8px 8px; }
            .bndd-row > div:nth-child(1) { flex: 2; }
            .bndd-row > div:nth-child(2) { flex: 2; }
            .bndd-row > div:nth-child(3) { flex: 1.5; }
            .bndd-row > div:nth-child(4) { flex: 1; }
            </style>
            """, unsafe_allow_html=True)
            
            # Header
            st.markdown("""
            <div class="bndd-header">
                <div>Facility</div>
                <div>License Number</div>
                <div>Expiration Date</div>
                <div>Actions</div>
            </div>
            """, unsafe_allow_html=True)
            
            # Rows with Streamlit columns for interactive elements
            for i, lic in enumerate(sorted_licenses):
                try:
                    exp_date = datetime.strptime(lic["expiration_date"], "%Y-%m-%d")
                    exp_display = exp_date.strftime("%b %d, %Y")
                    days_until = (exp_date - today).days
                    if days_until < 0:
                        exp_color = "#dc2626"  # Red - expired
                    elif days_until <= 30:
                        exp_color = "#ea580c"  # Orange - urgent
                    elif days_until <= 90:
                        exp_color = "#ca8a04"  # Yellow - warning
                    else:
                        exp_color = "#059669"  # Green - good
                except:
                    exp_display = lic.get("expiration_date", "N/A")
                    exp_color = "#64748b"
                
                col1, col2, col3, col4 = st.columns([2, 2, 1.5, 1])
                with col1:
                    st.markdown(f"**{lic['facility']}**")
                with col2:
                    st.text(lic.get("license_number", "N/A"))
                with col3:
                    st.markdown(f"<span style='color:{exp_color};font-weight:600;'>{exp_display}</span>", unsafe_allow_html=True)
                with col4:
                    if st.button("🔄 Renew", key=f"renew_{i}", use_container_width=True):
                        st.session_state[f"renewing_{i}"] = True
                        st.rerun()
                
                # Renewal form (appears when Renew is clicked)
                if st.session_state.get(f"renewing_{i}", False):
                    with st.form(f"renew_form_{i}"):
                        rcol1, rcol2, rcol3 = st.columns([2, 2, 1])
                        with rcol1:
                            try:
                                current_exp = datetime.strptime(lic["expiration_date"], "%Y-%m-%d").date()
                            except:
                                current_exp = datetime.now().date()
                            new_exp_date = st.date_input("New Expiration Date", value=current_exp, key=f"new_exp_{i}")
                        with rcol2:
                            st.write("")
                            st.write("")
                            submitted = st.form_submit_button("💾 Save", use_container_width=True)
                        with rcol3:
                            st.write("")
                            st.write("")
                            if st.form_submit_button("❌ Cancel", use_container_width=True):
                                st.session_state[f"renewing_{i}"] = False
                                st.rerun()
                        
                        if submitted:
                            # Find and update the license
                            for j, existing_lic in enumerate(bndd_licenses):
                                if existing_lic["facility"] == lic["facility"]:
                                    bndd_licenses[j]["expiration_date"] = new_exp_date.strftime("%Y-%m-%d")
                                    break
                            supa.save_bndd_licenses(bndd_licenses)
                            st.session_state.bndd_licenses = bndd_licenses
                            st.session_state[f"renewing_{i}"] = False
                            st.success(f"Updated expiration date for {lic['facility']}")
                            st.rerun()
            
            # Edit/Delete section
            with st.expander("✏️ Edit or Remove Licenses", expanded=False):
                for i, lic in enumerate(bndd_licenses):
                    with st.expander(f"**{lic['facility']}**"):
                        edit_col1, edit_col2, edit_col3 = st.columns([2, 2, 1])
                        with edit_col1:
                            edit_license_num = st.text_input(
                                "License Number",
                                value=lic.get("license_number", ""),
                                key=f"edit_lic_{i}"
                            )
                        with edit_col2:
                            try:
                                cur_date = datetime.strptime(lic.get("expiration_date", "2025-01-01"), "%Y-%m-%d").date()
                            except:
                                cur_date = datetime.now().date()
                            edit_exp = st.date_input("Expiration Date", value=cur_date, key=f"edit_exp_{i}")
                        with edit_col3:
                            st.write("")
                            if st.button("💾 Save", key=f"save_edit_{i}", use_container_width=True):
                                bndd_licenses[i]["license_number"] = edit_license_num.strip()
                                bndd_licenses[i]["expiration_date"] = edit_exp.strftime("%Y-%m-%d")
                                supa.save_bndd_licenses(bndd_licenses)
                                st.session_state.bndd_licenses = bndd_licenses
                                st.success("Saved!")
                                st.rerun()
                            if st.button("🗑️ Delete", key=f"del_lic_{i}", type="secondary", use_container_width=True):
                                del bndd_licenses[i]
                                supa.save_bndd_licenses(bndd_licenses)
                                st.session_state.bndd_licenses = bndd_licenses
                                st.rerun()
        else:
            st.info("No BNDD licenses tracked yet. Add your first license above.")
    
    # ============ TAB 3: Cubex Re-Stock ============
    with qa_tab3:
        st.markdown("### Cubex Re-Stock Tracking")
        st.caption("Track Cubex machine re-stock dates. Next re-stock is always 11 months after the last re-stock.")
        
        # Add New Cubex Entry
        with st.expander("➕ Add New Cubex Entry", expanded=False):
            with st.form("add_cubex_form", clear_on_submit=True):
                if facility_names_list:
                    col1, col2, col3 = st.columns([2, 2, 2])
                    with col1:
                        new_cubex_facility = st.selectbox("Facility", facility_names_list, key="cubex_fac_select")
                    with col2:
                        new_serial = st.text_input("Serial Number (optional)", placeholder="e.g., CBX-12345")
                    with col3:
                        new_restock_date = st.date_input("Re-Stock Date", key="cubex_restock_date")
                    
                    if st.form_submit_button("➕ Add Entry", use_container_width=True):
                        if new_cubex_facility:
                            # Calculate next restock (11 months from restock date)
                            restock_dt = datetime.combine(new_restock_date, datetime.min.time())
                            next_restock = restock_dt + relativedelta(months=11)
                            
                            cubex_restock.append({
                                "facility": new_cubex_facility,
                                "serial_number": new_serial.strip() if new_serial.strip() else "",
                                "restock_date": new_restock_date.strftime("%Y-%m-%d"),
                                "next_restock_due": next_restock.strftime("%Y-%m-%d"),
                            })
                            supa.save_cubex_restock(cubex_restock)
                            st.session_state.cubex_restock = cubex_restock
                            st.success(f"Added Cubex entry for {new_cubex_facility}")
                            st.rerun()
                        else:
                            st.warning("Please fill in all fields")
                else:
                    st.warning("No facilities found. Add facilities in Pharmacy Management first.")
                    st.form_submit_button("➕ Add Entry", disabled=True)
        
        # Display Cubex entries table with inline edit
        if cubex_restock:
            st.markdown(f"**{len(cubex_restock)} Cubex machines tracked**")
            
            # Sort alphabetically by facility name
            sorted_cubex = sorted(cubex_restock, key=lambda x: x.get("facility", "").lower())
            
            today = datetime.now()
            
            # Build display data for DataFrame
            display_rows = []
            for entry in sorted_cubex:
                try:
                    restock_dt = datetime.strptime(entry["restock_date"], "%Y-%m-%d")
                    next_due_dt = datetime.strptime(entry["next_restock_due"], "%Y-%m-%d")
                    days_until = (next_due_dt - today).days
                except:
                    restock_dt = None
                    next_due_dt = None
                    days_until = None
                
                display_rows.append({
                    "Facility": entry["facility"],
                    "Serial Number": entry.get("serial_number", "") or "—",
                    "Re-Stock Date": restock_dt.strftime("%b %d, %Y") if restock_dt else "N/A",
                    "Next Due": next_due_dt.strftime("%b %d, %Y") if next_due_dt else "N/A",
                    "Days": days_until if days_until is not None else 999,
                })
            
            # Create DataFrame
            df = pd.DataFrame(display_rows)
            
            # Style function for row colors based on Days column
            def color_cubex_rows(row):
                days = row["Days"]
                if days == 999 or days == "N/A":
                    return ["background-color: white"] * len(row)
                if days <= 30:
                    return ["background-color: #fecaca"] * len(row)  # Red
                elif days <= 60:
                    return ["background-color: #fef08a"] * len(row)  # Yellow
                elif days <= 90:
                    return ["background-color: #bbf7d0"] * len(row)  # Green
                return ["background-color: white"] * len(row)
            
            # Apply styling
            styled_df = df.style.apply(color_cubex_rows, axis=1)
            
            # Display the styled dataframe
            st.dataframe(styled_df, use_container_width=True, hide_index=True)
            
            st.write("")  # Spacing
            
            # Edit section below table
            with st.expander("✏️ Edit Entries", expanded=False):
                for idx, entry in enumerate(sorted_cubex):
                    actual_idx = next(j for j, e in enumerate(cubex_restock) if e["facility"] == entry["facility"] and e.get("serial_number") == entry.get("serial_number"))
                    
                    with st.expander(f"**{entry['facility']}** — {entry.get('serial_number', '') or 'No serial'}"):
                        ecol1, ecol2, ecol3 = st.columns([2, 2, 1])
                        with ecol1:
                            edit_serial = st.text_input("Serial Number", value=entry.get("serial_number", ""), key=f"edit_serial_{idx}")
                        with ecol2:
                            try:
                                cur_restock = datetime.strptime(entry.get("restock_date", "2025-01-01"), "%Y-%m-%d").date()
                            except:
                                cur_restock = datetime.now().date()
                            edit_restock = st.date_input("Re-Stock Date", value=cur_restock, key=f"edit_restock_{idx}")
                        with ecol3:
                            st.write("")
                            if st.button("💾 Save", key=f"save_cubex_{idx}", use_container_width=True):
                                restock_dt_new = datetime.combine(edit_restock, datetime.min.time())
                                next_restock = restock_dt_new + relativedelta(months=11)
                                
                                cubex_restock[actual_idx]["serial_number"] = edit_serial.strip()
                                cubex_restock[actual_idx]["restock_date"] = edit_restock.strftime("%Y-%m-%d")
                                cubex_restock[actual_idx]["next_restock_due"] = next_restock.strftime("%Y-%m-%d")
                                supa.save_cubex_restock(cubex_restock)
                                st.session_state.cubex_restock = cubex_restock
                                st.success("Saved!")
                                st.rerun()
                            if st.button("🗑️ Delete", key=f"del_cubex_{idx}", type="secondary", use_container_width=True):
                                del cubex_restock[actual_idx]
                                supa.save_cubex_restock(cubex_restock)
                                st.session_state.cubex_restock = cubex_restock
                                st.rerun()
        else:
            st.info("No Cubex machines tracked yet. Add your first entry above.")

# --- PAGE: Data Explorer ---
if current_page == "Data Explorer":
    st.markdown("### Demo data explorer")
    st.caption("Local prototype data Turner can edit immediately.")
    st.code(DATA_FILE.read_text(), language="json")

# --- PAGE: User Management (admin only) ---
if current_page == "User Management" and is_admin_user():
    st.markdown("### 👥 User Management")
    st.caption("Add, edit, or remove users. Manage permissions.")
    
    # Load current config (from Supabase + local merged)
    user_config = load_merged_config()
    
    users = user_config.get("credentials", {}).get("usernames", {})
    
    # Show all users in a table at the top
    st.markdown("#### All Users")
    if users:
        user_rows = []
        for username, user_data in sorted(users.items()):
            user_rows.append({
                "Username": username,
                "Display Name": user_data.get("name", "N/A"),
                "Role": user_data.get("role", "user").capitalize(),
            })
        st.dataframe(pd.DataFrame(user_rows), use_container_width=True, hide_index=True)
    else:
        st.info("No users configured.")
    
    st.divider()
    
    # Add new user
    st.markdown("#### Add New User")
    with st.form("add_user_form", clear_on_submit=True):
        col1, col2 = st.columns(2)
        with col1:
            new_username = st.text_input("Username", placeholder="e.g., jsmith")
            new_display = st.text_input("Display Name", placeholder="e.g., John Smith")
        with col2:
            new_password = st.text_input("Password", type="password")
            new_role = st.selectbox("Role", ["user", "admin"])
        
        if st.form_submit_button("➕ Add User", use_container_width=True):
            if new_username and new_password and new_display:
                if len(new_password) < 6:
                    st.error("Password must be at least 6 characters")
                elif new_username in users:
                    st.error(f"Username '{new_username}' already exists")
                else:
                    import bcrypt
                    hashed = bcrypt.hashpw(new_password.encode(), bcrypt.gensalt()).decode()
                    user_config["credentials"]["usernames"][new_username] = {
                        "name": new_display,
                        "password": hashed,
                        "role": new_role
                    }
                    save_config(user_config)
                    st.success(f"Added user '{new_username}'")
                    st.rerun()
            else:
                st.warning("Please fill in all fields")
    
    st.divider()
    
    # List existing users
    st.markdown("#### Current Users")
    for username, user_data in users.items():
        with st.expander(f"**{user_data.get('name', username)}** (@{username}) — {user_data.get('role', 'user')}"):
            col1, col2, col3 = st.columns([2, 1, 1])
            
            with col1:
                # Change password
                new_pwd = st.text_input("New Password", type="password", key=f"pwd_{username}", placeholder="Leave blank to keep")
                if new_pwd:
                    if st.button("Update Password", key=f"update_pwd_{username}"):
                        if len(new_pwd) >= 6:
                            import bcrypt
                            hashed = bcrypt.hashpw(new_pwd.encode(), bcrypt.gensalt()).decode()
                            user_config["credentials"]["usernames"][username]["password"] = hashed
                            save_config(user_config)
                            st.success("Password updated")
                            st.rerun()
                        else:
                            st.error("Min 6 characters")
            
            with col2:
                # Change role
                current_role = user_data.get("role", "user")
                new_role_sel = st.selectbox("Role", ["user", "admin"], 
                                       index=0 if current_role == "user" else 1,
                                       key=f"role_{username}")
                if new_role_sel != current_role:
                    if st.button("Update Role", key=f"update_role_{username}"):
                        user_config["credentials"]["usernames"][username]["role"] = new_role_sel
                        save_config(user_config)
                        st.success(f"Role updated to {new_role_sel}")
                        st.rerun()
            
            with col3:
                # Delete user (can't delete yourself)
                if username != st.session_state.get("username"):
                    if st.button("🗑️ Delete", key=f"del_{username}", type="secondary"):
                        del user_config["credentials"]["usernames"][username]
                        save_config(user_config)
                        st.success(f"Deleted user '{username}'")
                        st.rerun()
                else:
                    st.caption("(You)")
    
    st.divider()
    
    # --- Permissions Management ---
    st.markdown("#### 🔐 Page Permissions")
    st.caption("Control which pages each role or user can access.")
    
    permissions = user_config.get("permissions", {"roles": {"admin": ["all"], "user": []}, "users": {}})
    
    # Role permissions
    st.markdown("##### Role Defaults")
    role_col1, role_col2 = st.columns(2)
    
    with role_col1:
        st.markdown("**Admin Role**")
        st.caption("Admins have access to all pages by default.")
    
    with role_col2:
        st.markdown("**User Role**")
        current_user_perms = permissions.get("roles", {}).get("user", [])
        
        # Checkboxes for each page
        new_user_perms = []
        for page in ALL_PAGES:
            if page == "User Management":
                continue  # Never give regular users access to user management
            checked = st.checkbox(page, value=(page in current_user_perms), key=f"perm_user_{page}")
            if checked:
                new_user_perms.append(page)
        
        if new_user_perms != current_user_perms:
            if st.button("Save User Role Permissions"):
                if "permissions" not in user_config:
                    user_config["permissions"] = {"roles": {}, "users": {}}
                if "roles" not in user_config["permissions"]:
                    user_config["permissions"]["roles"] = {}
                user_config["permissions"]["roles"]["user"] = new_user_perms
                save_config(user_config)
                st.success("User role permissions updated!")
                st.rerun()
    
    st.divider()
    
    # Per-user permission overrides
    st.markdown("##### Per-User Overrides")
    st.caption("Override role permissions for specific users.")
    
    user_overrides = permissions.get("users", {})
    
    for username, user_data in users.items():
        user_override = user_overrides.get(username)
        role = user_data.get("role", "user")
        
        with st.expander(f"**{user_data.get('name', username)}** (@{username}) — {role}"):
            has_override = user_override is not None
            enable_override = st.checkbox("Enable custom permissions", value=has_override, key=f"override_enable_{username}")
            
            if enable_override:
                current_perms = user_override if user_override else []
                new_perms = []
                
                cols = st.columns(3)
                for i, page in enumerate(ALL_PAGES):
                    if page == "User Management" and role != "admin":
                        continue
                    with cols[i % 3]:
                        checked = st.checkbox(page, value=(page in current_perms), key=f"perm_{username}_{page}")
                        if checked:
                            new_perms.append(page)
                
                if st.button(f"Save {username}'s permissions", key=f"save_perm_{username}"):
                    if "permissions" not in user_config:
                        user_config["permissions"] = {"roles": {}, "users": {}}
                    if "users" not in user_config["permissions"]:
                        user_config["permissions"]["users"] = {}
                    user_config["permissions"]["users"][username] = new_perms
                    save_config(user_config)
                    st.success(f"Permissions for {username} updated!")
                    st.rerun()
            else:
                if has_override:
                    if st.button(f"Remove override (use role defaults)", key=f"remove_override_{username}"):
                        del user_config["permissions"]["users"][username]
                        save_config(user_config)
                        st.success(f"Override removed for {username}")
                        st.rerun()
                else:
                    st.caption(f"Using {role} role defaults")

st.divider()
last_updated = datetime.fromtimestamp(DATA_FILE.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S")
st.caption(f"Ozark LTC Rx Ops prototype ready · Data source: {DATA_FILE} · Last data update: {last_updated}")
